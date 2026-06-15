"""Importe les Cours d'Appel + Tribunaux de Première Instance depuis le XLSX.

Source : media/المحاكم الابتدائية مع الاستئناف.xlsx

Structure attendue (colonnes K..N) :
    K (idx 10) = ر.ت (ordre interne de la PI dans le groupe)
    L (idx 11) = nom de la PI (Tribunal de Première Instance / centre judiciaire)
    M (idx 12) = code numérique de la CA parente (0..20)
    N (idx 13) = nom de la CA parente

Les groupes de PI sont séparés par une ligne vide. La première occurrence d'un
code CA dans la colonne M permet de créer la CA elle-même.

Usage :
    python manage.py import_juridictions_xlsx
"""
from __future__ import annotations

from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

from avocat_app.models import Juridiction, TypeJuridiction


XLSX_REL_PATH = "media/المحاكم الابتدائية مع الاستئناف.xlsx"
# Colonnes (1-indexées en openpyxl → 0-indexées dans iter_rows values_only)
COL_ORDRE_PI = 10
COL_NOM_PI = 11
COL_CODE_CA = 12
COL_NOM_CA = 13


def _normalize(s: str | None) -> str:
    if s is None:
        return ""
    return str(s).replace("\xa0", " ").strip()


def _ensure_type(code: str, libelle_ar: str, libelle_fr: str, niveau: str) -> TypeJuridiction:
    obj, _ = TypeJuridiction.objects.get_or_create(
        code_type=code,
        defaults={
            "libelle": libelle_ar,
            "libelle_fr": libelle_fr,
            "niveau": niveau,
            "description": libelle_ar,
        },
    )
    return obj


class Command(BaseCommand):
    help = "Importe Cours d'Appel + Tribunaux de 1ère Instance depuis le XLSX."

    def add_arguments(self, parser):
        parser.add_argument("--path", default=None,
                            help="Chemin du XLSX (par défaut: media/المحاكم الابتدائية مع الاستئناف.xlsx)")
        parser.add_argument("--dry-run", action="store_true",
                            help="Affiche sans écrire en base.")

    @transaction.atomic
    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stdout.write(self.style.ERROR("openpyxl manquant. pip install openpyxl"))
            return

        path = Path(options.get("path") or (Path(settings.BASE_DIR) / XLSX_REL_PATH))
        if not path.exists():
            self.stdout.write(self.style.ERROR(f"Fichier introuvable : {path}"))
            return

        dry = options["dry_run"]

        # Types nécessaires
        type_ca = _ensure_type("CA", "محكمة الاستئناف", "Cour d'Appel", "appel")
        type_pi = _ensure_type("TPI", "المحكمة الابتدائية", "Tribunal de Première Instance", "premiere_instance")
        type_cj = _ensure_type("CJ", "المركز القضائي", "Centre Judiciaire", "premiere_instance")
        type_qf = _ensure_type("QAF", "قسم قضاء الأسرة", "Section Famille", "premiere_instance")

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active

        # Étape 1 : collecter toutes les (code_CA, nom_CA) → créer les CA d'abord.
        ca_map: dict[int, Juridiction] = {}
        pi_rows: list[tuple[int, str, int, str]] = []  # (ordre_pi, nom_pi, code_ca, nom_ca)

        for row in ws.iter_rows(values_only=True):
            if len(row) <= COL_NOM_CA:
                continue
            ordre_pi = row[COL_ORDRE_PI]
            nom_pi = _normalize(row[COL_NOM_PI])
            code_ca = row[COL_CODE_CA]
            nom_ca = _normalize(row[COL_NOM_CA])
            if code_ca is None or not nom_ca or nom_ca.startswith("ر.ت") or "محكمة النقض" in nom_ca:
                continue
            try:
                code_ca_int = int(code_ca)
            except (TypeError, ValueError):
                continue
            if not isinstance(ordre_pi, int) or not nom_pi:
                continue
            pi_rows.append((ordre_pi, nom_pi, code_ca_int, nom_ca))

        # Créer les CA uniques
        ca_unique: dict[int, str] = {}
        for _, _, code_ca, nom_ca in pi_rows:
            ca_unique.setdefault(code_ca, nom_ca)

        self.stdout.write(f"→ {len(ca_unique)} Cour(s) d'Appel à créer/mettre à jour")
        for code_ca, nom_ca in sorted(ca_unique.items()):
            code_str = f"CA{code_ca:02d}"
            if dry:
                self.stdout.write(f"  [DRY] CA #{code_ca} ({code_str}) : {nom_ca}")
                continue
            ca, created = Juridiction.objects.update_or_create(
                code=code_str,
                defaults={
                    "nomtribunal_ar": nom_ca,
                    "nomtribunal_fr": f"Cour d'Appel #{code_ca}",
                    "type": type_ca,
                    "TribunalParent": None,
                },
            )
            ca_map[code_ca] = ca
            self.stdout.write(f"  ✓ CA #{code_ca} {'créée' if created else 'mise à jour'} : {nom_ca}")

        # Étape 2 : créer les PI rattachées à leur CA parente
        self.stdout.write(f"\n→ {len(pi_rows)} Tribunaux de 1ère Instance à créer/mettre à jour")
        created_pi = updated_pi = 0
        for ordre_pi, nom_pi, code_ca, _ in pi_rows:
            # Choisir le type selon le nom
            if "قسم قضاء الأسرة" in nom_pi:
                tj = type_qf
            elif "المركز القضائي" in nom_pi:
                tj = type_cj
            else:
                tj = type_pi
            code_pi = f"PI{code_ca:02d}-{ordre_pi:02d}"
            if dry:
                self.stdout.write(f"  [DRY] {code_pi} → {nom_pi} (parent: CA#{code_ca})")
                continue
            parent = ca_map.get(code_ca)
            obj, was_created = Juridiction.objects.update_or_create(
                code=code_pi,
                defaults={
                    "nomtribunal_ar": nom_pi,
                    "nomtribunal_fr": "",
                    "type": tj,
                    "TribunalParent": parent,
                },
            )
            if was_created:
                created_pi += 1
            else:
                updated_pi += 1

        if dry:
            self.stdout.write(self.style.WARNING("\n[DRY-RUN] Aucune écriture effectuée."))
        else:
            self.stdout.write(self.style.SUCCESS(
                f"\n✓ Terminé : {len(ca_map)} CA, {created_pi} PI créées, {updated_pi} PI mises à jour."
            ))
