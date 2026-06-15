"""Importe les codes d'affaires (~540 codes) depuis le XLSX رموز المحاكم.xlsx.

Structure attendue (sheet1) :
    Col A : N°  (numéro ligne)
    Col B : رموز الملفات     (catégorie globale - civil/pénal/admin/commercial)
    Col C : code_type1        (code 4 chiffres - groupe / chambre, ex 1100, 1200…)
    Col D : libelle           (libellé de la chambre, ex "مؤسسة الرئيس وغرفة المشورة")
    Col E : sous_type         (ابتدائي / استئنافي / None)
    Col F : Code_TypeAffaire  (code 4 chiffres affaire, ex 1101, 1102…)
    Col G : code affaire      (libellé affaire, ex "الاستعجالي")

Usage :
    python manage.py import_codes_affaires_xlsx
"""
from __future__ import annotations

from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

from avocat_app.models import CodeCategorieAffaire


XLSX_REL_PATH = "media/رموز المحاكم.xlsx"


# Mappe le libellé de la catégorie globale du XLSX vers notre clé interne.
CATEGORIE_GLOBALE_MAP = {
    "رموز الملفات بالمحاكم المدنية":  "civil",
    "رموز الملفات الجنحية":            "penal",
    "رموز الملفات بالمحاكم الإدارية": "admin",
    "رموز الملفات بالمحاكم التجارية": "commercial",
}


# Domaine d'arrière-plan : déduit de la catégorie globale, sauf override par chambre.
# (Le champ `domaine` du modèle pré-existe avec 11 valeurs pour compatibilité.)
GLOBAL_TO_DOMAINE = {
    "civil":      "civil",
    "penal":      "penal",
    "admin":      "administratif",
    "commercial": "commercial",
}


# Override par code_type (chambre) pour affiner le domaine quand on connaît.
CODE_TYPE_DOMAINE_OVERRIDE = {
    "1300": "civil",         # الأكرية
    "1400": "immobilier",
    "1500": "social",
    "1600": "famille",
    "1700": "proximite",
    "6100": "execution",
    "6150": "execution",
    "6200": "execution",
    "6250": "execution",
    "6300": "execution",
    "6350": "execution",
    "6500": "notification",
    "6600": "notification",
    "6700": "notification",
    "6800": "notification",
    "7100": "administratif",
    "7200": "administratif",
}


# Mappe sous_type vers le champ `niveau` historique (premiere_instance / appel).
SOUS_TYPE_TO_NIVEAU = {
    "ابتدائي":   "premiere_instance",
    "استئنافي":  "appel",
}


def _normalize(s) -> str:
    if s is None:
        return ""
    return str(s).replace("\xa0", " ").strip()


class Command(BaseCommand):
    help = "Importe les codes d'affaires depuis رموز المحاكم.xlsx."

    def add_arguments(self, parser):
        parser.add_argument("--path", default=None,
                            help="Chemin du XLSX (par défaut: media/رموز المحاكم.xlsx)")
        parser.add_argument("--dry-run", action="store_true",
                            help="Affiche sans écrire en base.")

    @transaction.atomic
    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stdout.write(self.style.ERROR("openpyxl manquant. pip install openpyxl"))
            return

        path = Path(options.get("path") or (Path(settings.BASE_DIR) / XLSX_REL_PATH))
        if not path.exists():
            self.stdout.write(self.style.ERROR(f"Fichier introuvable : {path}"))
            return

        dry = options["dry_run"]

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        # Première feuille (ورقة1)
        ws = wb[wb.sheetnames[0]]

        created = updated = skipped = 0
        seen_codes: set[str] = set()

        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 1:
                continue  # header
            if not row or len(row) < 7:
                continue
            categorie_xlsx = _normalize(row[1])
            code_type = _normalize(row[2])
            type_libelle = _normalize(row[3])
            sous_type = _normalize(row[4])
            code_affaire = _normalize(row[5])
            libelle = _normalize(row[6])

            if not code_affaire or not libelle:
                skipped += 1
                continue
            if code_affaire in seen_codes:
                skipped += 1
                continue
            seen_codes.add(code_affaire)

            categorie_globale = CATEGORIE_GLOBALE_MAP.get(categorie_xlsx, "")
            domaine = CODE_TYPE_DOMAINE_OVERRIDE.get(code_type) or \
                      GLOBAL_TO_DOMAINE.get(categorie_globale, "civil")
            niveau = SOUS_TYPE_TO_NIVEAU.get(sous_type, "premiere_instance")

            if dry:
                self.stdout.write(
                    f"  [DRY] {code_affaire} | {libelle} | type={code_type} ({type_libelle}) "
                    f"| sous={sous_type or '—'} | dom={domaine} | glob={categorie_globale}"
                )
                continue

            _, was_created = CodeCategorieAffaire.objects.update_or_create(
                code=code_affaire,
                defaults={
                    "libelle": libelle,
                    "domaine": domaine,
                    "niveau": niveau,
                    "code_type": code_type,
                    "type_libelle": type_libelle,
                    "sous_type": sous_type,
                    "categorie_globale": categorie_globale,
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

        if dry:
            self.stdout.write(self.style.WARNING(
                f"\n[DRY-RUN] {len(seen_codes)} codes uniques détectés, {skipped} sautés."
            ))
        else:
            self.stdout.write(self.style.SUCCESS(
                f"\n✓ Terminé : {created} créés, {updated} mis à jour, {skipped} sautés."
            ))
