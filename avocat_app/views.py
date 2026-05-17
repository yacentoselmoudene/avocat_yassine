from __future__ import annotations

from typing import Any, Dict
from datetime import timedelta
import calendar

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db.models import Q, Sum
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.formats import date_format
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView

from .views_mixins import (
    NoPostOnReadOnlyMixin, HTMXViewMixin, SoftDeleteQuerysetMixin,
    ModalCreateView, ModalUpdateView, ModalDeleteView, HTMXModalFormMixin,
    UIPermRequiredMixin,)

from .models import (
    Juridiction, Avocat, Affaire, Partie, AffairePartie, AffaireAvocat,
    Audience, Mesure, Expertise, Decision, Notification, VoieDeRecours,
    Execution, Depense, Recette, PieceJointe, Utilisateur, Tache, Alerte, Expert, Barreau,
    Avertissement, PhaseAffaire, DocumentRequirement, MahakimSyncResult,
    ContumaceRecord, WhatsAppTemplate, WhatsAppMessage, DecisionAnalysis,
)

from .filters import AffaireFilter, DepenseFilter, RecetteFilter, PieceJointeFilter, AudienceFilter

from .forms import (
    JuridictionForm, AvocatForm, AffaireForm, PartieForm, AffairePartieForm, AffaireAvocatForm,
    AudienceForm, MesureForm, ExpertiseForm, DecisionForm, NotificationForm, VoieDeRecoursForm,
    ExecutionForm, DepenseForm, RecetteForm, PieceJointeForm, UtilisateurForm, TacheForm, AlerteForm, ExpertForm,
    BarreauForm, AvertissementForm,
)

# -------------------------------------------------------------
# Utilitaires communs
# -------------------------------------------------------------
class SecureBase(LoginRequiredMixin, PermissionRequiredMixin):
    login_url = reverse_lazy('authui:login')
    raise_exception = False
    permission_required = ''

    def success_json(self, message=None, **payload):
        """
        JSON نجاح مرن يقبل أي مفاتيح إضافية:
        - redirect: URL
        - refreshTarget: CSS selector
        - refreshUrl: URL
        - html: HTML للاستخدام في Toast
        - message: نص مختصر
        """
        data = {"ok": True}
        if message:
            data["message"] = message
        # اندمج أي مفاتيح إضافية (refreshTarget/refreshUrl/redirect/html/...)
        data.update(payload)
        return JsonResponse(data)

# -------------------------------------------------------------
# Aide pour retrouver l'affaire liée à une étape (audience/mesure/…)
# -------------------------------------------------------------
def _affaire_pk_from_step(obj) -> int | None:
    if obj is None:
        return None
    if isinstance(obj, Audience):
        return obj.affaire_id
    if isinstance(obj, Mesure):
        return obj.audience.affaire_id if obj.audience_id else None
    if isinstance(obj, Expertise):
        return obj.affaire_id
    if isinstance(obj, Decision):
        return obj.affaire_id
    if isinstance(obj, Notification):
        return obj.decision.affaire_id if obj.decision_id else None
    if isinstance(obj, VoieDeRecours):
        return obj.decision.affaire_id if obj.decision_id else None
    if isinstance(obj, Execution):
        return obj.decision.affaire_id if obj.decision_id else None
    if isinstance(obj, Avertissement):
        return obj.affaire_id
    return None




class SearchListMixin:
    """Mixin générique pour filtrer via ?q= sur un ListView."""
    search_param = "q"
    search_fields: list[str] = []

    def get_search_query(self) -> str:
        return (self.request.GET.get(self.search_param) or "").strip()

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.get_search_query()
        if q and self.search_fields:
            cond = Q()
            for f in self.search_fields:
                cond |= Q(**{f"{f}__icontains": q})
            qs = qs.filter(cond)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["query"] = self.get_search_query()
        return ctx


class DjangoFilterListMixin:
    """Mixin pour intégrer django-filter dans un ListView."""
    filterset_class = None

    def get_queryset(self):
        qs = super().get_queryset()
        if self.filterset_class:
            self.filterset = self.filterset_class(self.request.GET, queryset=qs)
            return self.filterset.qs
        self.filterset = None
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        fs = getattr(self, "filterset", None)
        ctx["filter"] = fs
        count = 0
        if fs and fs.form.is_bound:
            skip = {"q", "page", "view"}
            for name, val in fs.form.cleaned_data.items():
                if name not in skip and val not in (None, "", []):
                    count += 1
        ctx["active_filter_count"] = count
        return ctx


class HTMXPartialListMixin:
    """Si la requête est HTMX, ne renvoie que le partial du tableau."""
    partial_template_name: str = None  # à définir sur la vue

    def render_to_response(self, context, **response_kwargs):
        if self.request.headers.get("HX-Request") and self.partial_template_name:
            return render(self.request, self.partial_template_name, context)
        return super().render_to_response(context, **response_kwargs)


# =============================================================
# POPUPS PERSONNE (Partie / Expert / Avocat / Utilisateur)
# =============================================================

class _PersonPopupBase(SecureBase, HTMXViewMixin, DetailView):
    """Base pour les popups read-only de personnes."""

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        ctx = self.build_popup_context()
        return self.render_modal("modals/_person_popup.html", ctx)

    def build_popup_context(self):
        raise NotImplementedError


class PartiePopup(_PersonPopupBase):
    model = Partie
    permission_required = "cabinet.view_partie"

    def build_popup_context(self):
        o = self.object
        return {
            "person_name": o.nom_complet,
            "person_badge": o.get_type_partie_display(),
            "icon": "bi-person-fill",
            "detail_url": o.get_absolute_url(),
            "person_fields": [
                {"label": "رقم التعريف", "value": o.cin_ou_rc},
                {"label": "العنوان", "value": o.adresse},
                {"label": "الهاتف", "value": o.telephone},
                {"label": "البريد", "value": o.email},
                {"label": "الممثل القانوني", "value": o.representant_legal},
                {"label": "المحامي", "value": str(o.avocat) if o.avocat else None},
            ],
        }


class ExpertPopup(_PersonPopupBase):
    model = Expert
    permission_required = "cabinet.view_expert"

    def build_popup_context(self):
        o = self.object
        return {
            "person_name": o.nom_complet,
            "person_badge": o.specialite,
            "icon": "bi-mortarboard-fill",
            "detail_url": o.get_absolute_url(),
            "person_fields": [
                {"label": "التخصص", "value": o.specialite},
                {"label": "الهاتف", "value": o.telephone},
                {"label": "البريد", "value": o.email},
                {"label": "العنوان", "value": o.adresse},
            ],
        }


class AvocatPopup(_PersonPopupBase):
    model = Avocat
    permission_required = "cabinet.view_avocat"

    def build_popup_context(self):
        o = self.object
        return {
            "person_name": o.nom,
            "person_badge": str(o.barreau),
            "icon": "bi-briefcase-fill",
            "detail_url": o.get_absolute_url(),
            "person_fields": [
                {"label": "الهيئة", "value": str(o.barreau)},
                {"label": "الهاتف", "value": o.telephone},
                {"label": "البريد", "value": o.email},
                {"label": "التعريفة/ساعة", "value": f"{o.taux_horaire} د.م." if o.taux_horaire else None},
            ],
        }


class UtilisateurPopup(_PersonPopupBase):
    model = Utilisateur
    permission_required = "cabinet.view_utilisateur"

    def build_popup_context(self):
        o = self.object
        return {
            "person_name": o.nom_complet,
            "person_badge": str(o.role),
            "icon": "bi-person-badge-fill",
            "detail_url": o.get_absolute_url(),
            "person_fields": [
                {"label": "الدور", "value": str(o.role)},
                {"label": "البريد", "value": o.email},
                {"label": "الهاتف", "value": o.telephone},
                {"label": "نشط", "value": "نعم" if o.actif else "لا"},
            ],
        }


# =============================================================
# DASHBOARD
# =============================================================

class DashboardView(SecureBase, TemplateView):
    template_name = "dashboard/index.html"
    permission_required = "cabinet.view_affaire"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = timezone.localdate()
        upcoming = today + timedelta(days=14)
        deadline_horizon = today + timedelta(days=7)

        ctx["stats"] = {
            "affaires_total": Affaire.objects.count(),
            "audiences_a_venir": Audience.objects.filter(date_audience__range=(today, upcoming)).count(),
        }

        ctx["next_audiences"] = (
            Audience.objects.select_related("affaire")
            .filter(date_audience__gte=today).order_by("date_audience")[:6]
        )
        ctx["recent_affaires"] = (
            Affaire.objects.select_related("juridiction")
            .order_by("-date_ouverture")[:6]
        )
        # Finance navigation: ?fy=year&fm=month
        ARABIC_MONTHS = {
            1: "يناير", 2: "فبراير", 3: "مارس", 4: "أبريل",
            5: "ماي", 6: "يونيو", 7: "يوليوز", 8: "غشت",
            9: "شتنبر", 10: "أكتوبر", 11: "نونبر", 12: "دجنبر",
        }
        try:
            fin_year = int(self.request.GET.get("fy", today.year))
            fin_month = int(self.request.GET.get("fm", today.month))
            if not (1 <= fin_month <= 12):
                fin_year, fin_month = today.year, today.month
        except (ValueError, TypeError):
            fin_year, fin_month = today.year, today.month

        fin_is_current = (fin_year == today.year and fin_month == today.month)

        # Prev/next month calculation
        if fin_month == 1:
            prev_y, prev_m = fin_year - 1, 12
        else:
            prev_y, prev_m = fin_year, fin_month - 1
        if fin_month == 12:
            next_y, next_m = fin_year + 1, 1
        else:
            next_y, next_m = fin_year, fin_month + 1

        try:
            dep = Depense.objects.filter(
                date_depense__year=fin_year, date_depense__month=fin_month
            ).aggregate(Sum("montant"))["montant__sum"] or 0
            rec = Recette.objects.filter(
                date_recette__year=fin_year, date_recette__month=fin_month
            ).aggregate(Sum("montant"))["montant__sum"] or 0
            ctx["finance"] = {
                "depenses_mois": dep,
                "recettes_mois": rec,
                "net": rec - dep,
            }
        except Exception:
            ctx["finance"] = {"depenses_mois": 0, "recettes_mois": 0, "net": 0}

        ctx["fin_year"] = fin_year
        ctx["fin_month"] = fin_month
        ctx["fin_month_label"] = ARABIC_MONTHS.get(fin_month, "")
        ctx["fin_is_current"] = fin_is_current
        ctx["fin_prev_y"] = prev_y
        ctx["fin_prev_m"] = prev_m
        ctx["fin_next_y"] = next_y
        ctx["fin_next_m"] = next_m

        # Approaching deadlines (avertissements + recours within 7 days)
        approaching_deadlines = []
        for av in Avertissement.objects.filter(
            date_echeance__range=(today, deadline_horizon),
            resultat="en_attente"
        ).select_related("affaire", "type_avertissement")[:10]:
            approaching_deadlines.append({
                "type": "إنذار",
                "title": str(av.type_avertissement),
                "affaire": av.affaire.reference_interne,
                "affaire_pk": av.affaire_id,
                "echeance": av.date_echeance,
                "jours": av.jours_restants,
            })
        for r in VoieDeRecours.objects.filter(
            date_echeance_recours__range=(today, deadline_horizon),
        ).select_related("decision__affaire", "type_recours")[:10]:
            approaching_deadlines.append({
                "type": "طعن",
                "title": str(r.type_recours),
                "affaire": r.decision.affaire.reference_interne,
                "affaire_pk": r.decision.affaire_id,
                "echeance": r.date_echeance_recours,
                "jours": r.jours_restants_recours,
            })
        approaching_deadlines.sort(key=lambda x: x["echeance"])
        ctx["approaching_deadlines"] = approaching_deadlines

        # Phase distribution
        from django.db.models import Count
        phase_dist = (
            Affaire.objects.values("phase")
            .annotate(count=Count("id"))
            .order_by("phase")
        )
        phase_labels = dict(PhaseAffaire.choices)
        ctx["phase_distribution"] = [
            {"phase": phase_labels.get(p["phase"], p["phase"]), "phase_code": p["phase"], "count": p["count"]}
            for p in phase_dist
        ]

        # Widget "ما عليّ اليوم" (Today's actionable items)
        now = timezone.now()
        today_start = timezone.make_aware(
            timezone.datetime.combine(today, timezone.datetime.min.time())
        ) if timezone.is_naive(now) else now.replace(hour=0, minute=0, second=0, microsecond=0)
        tomorrow_start = today_start + timedelta(days=1)
        deadline_24h = today + timedelta(days=1)

        today_audiences = (
            Audience.objects.select_related("affaire", "affaire__juridiction", "type_audience")
            .filter(date_audience__gte=today_start, date_audience__lt=tomorrow_start)
            .order_by("date_audience")
        )

        urgent_avertissements = (
            Avertissement.objects.select_related("affaire", "type_avertissement")
            .filter(date_echeance__range=(today, deadline_24h), resultat="en_attente")
            .order_by("date_echeance")
        )

        urgent_recours = (
            VoieDeRecours.objects.select_related("decision__affaire", "type_recours")
            .filter(date_echeance_recours__range=(today, deadline_24h))
            .order_by("date_echeance_recours")
        )

        try:
            today_taches = list(
                Tache.objects.select_related("statut", "assigne_a")
                .filter(echeance__date=today)
                .order_by("echeance")[:5]
            )
        except Exception:
            today_taches = []

        ctx["today_widget"] = {
            "audiences": today_audiences,
            "avertissements": urgent_avertissements,
            "recours": urgent_recours,
            "taches": today_taches,
            "total_count": (
                today_audiences.count()
                + urgent_avertissements.count()
                + urgent_recours.count()
                + len(today_taches)
            ),
        }

        return ctx


# =============================================================
# RECHERCHE GLOBALE (HTMX autocomplete)
# =============================================================

@login_required(login_url=reverse_lazy('authui:login'))
def global_search(request: HttpRequest) -> HttpResponse:
    """Recherche globale multi-modèles renvoyant un dropdown HTML pour HTMX."""
    q = (request.GET.get("q") or "").strip()
    ctx = {"q": q, "groups": [], "too_short": False}

    if len(q) < 2:
        ctx["too_short"] = True
        return render(request, "_partials/_global_search_results.html", ctx)

    LIMIT = 5

    affaires = (
        Affaire.objects.select_related("juridiction", "type_affaire")
        .filter(
            Q(reference_interne__icontains=q)
            | Q(reference_tribunal__icontains=q)
            | Q(numero_dossier__icontains=q)
            | Q(objet__icontains=q)
        )[:LIMIT]
    )
    parties = (
        Partie.objects.filter(
            Q(nom_complet__icontains=q)
            | Q(telephone__icontains=q)
            | Q(cin_ou_rc__icontains=q)
            | Q(email__icontains=q)
        )[:LIMIT]
    )
    juridictions = (
        Juridiction.objects.filter(
            Q(nomtribunal_ar__icontains=q)
            | Q(nomtribunal_fr__icontains=q)
            | Q(villetribunal_ar__icontains=q)
            | Q(villetribunal_fr__icontains=q)
            | Q(code__icontains=q)
        )[:LIMIT]
    )
    avocats = (
        Avocat.objects.filter(
            Q(nom__icontains=q) | Q(telephone__icontains=q) | Q(email__icontains=q)
        )[:LIMIT]
    )

    groups = []
    if affaires:
        groups.append({
            "label": "القضايا", "icon": "bi-folder2",
            "items": [{
                "title": a.reference_interne,
                "subtitle": f"{a.type_affaire} — {a.juridiction.nomtribunal_ar or a.juridiction.nomtribunal_fr or ''}",
                "url": a.get_absolute_url(),
            } for a in affaires],
        })
    if parties:
        groups.append({
            "label": "الأطراف", "icon": "bi-person",
            "items": [{
                "title": p.nom_complet,
                "subtitle": f"{p.get_type_partie_display()} — {p.telephone or p.cin_ou_rc or p.email or ''}",
                "url": p.get_absolute_url(),
            } for p in parties],
        })
    if juridictions:
        groups.append({
            "label": "المحاكم", "icon": "bi-building",
            "items": [{
                "title": j.nomtribunal_ar or j.nomtribunal_fr or j.code,
                "subtitle": j.villetribunal_ar or j.villetribunal_fr or "",
                "url": j.get_absolute_url(),
            } for j in juridictions],
        })
    if avocats:
        groups.append({
            "label": "المحامون", "icon": "bi-person-badge",
            "items": [{
                "title": a.nom,
                "subtitle": a.telephone or a.email or "",
                "url": a.get_absolute_url(),
            } for a in avocats],
        })

    ctx["groups"] = groups
    return render(request, "_partials/_global_search_results.html", ctx)


# =============================================================
# AFFAIRES
# =============================================================

class AffaireList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, DjangoFilterListMixin, HTMXPartialListMixin, ListView):
    model = Affaire
    template_name = "avocat/affaire_list.html"
    permission_required = "cabinet.view_affaire"
    paginate_by = 20
    filterset_class = AffaireFilter

    search_fields = [
        "reference_interne", "reference_tribunal", "objet",
        "juridiction__nomtribunal_ar", "juridiction__nomtribunal_fr",
        "type_affaire__libelle", "avocat_responsable__nom",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("type_affaire", "statut_affaire", "juridiction", "avocat_responsable").order_by("-date_ouverture")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["view_mode"] = self.request.GET.get("view", "table")
        return ctx


class AffaireDetail(SecureBase, NoPostOnReadOnlyMixin, DetailView):
    model = Affaire
    template_name = "avocat/affaire_detail.html"
    permission_required = "cabinet.view_affaire"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        affaire = self.object
        phase = affaire.phase
        has_decision = Decision.objects.filter(affaire=affaire).exists()

        # Phase-aware workflow locks
        ctx["can_add_avertissement"] = phase in (PhaseAffaire.PRELIMINAIRE,)
        ctx["can_add_audience"] = phase in (PhaseAffaire.PRELIMINAIRE, PhaseAffaire.PREMIERE_INSTANCE, PhaseAffaire.APPEL, PhaseAffaire.CASSATION)
        ctx["can_add_decision"] = phase in (PhaseAffaire.PREMIERE_INSTANCE, PhaseAffaire.APPEL, PhaseAffaire.CASSATION)
        ctx["can_add_recours"] = has_decision and phase in (PhaseAffaire.PREMIERE_INSTANCE, PhaseAffaire.APPEL, PhaseAffaire.CASSATION)
        ctx["can_add_execution"] = has_decision and phase in (PhaseAffaire.PREMIERE_INSTANCE, PhaseAffaire.APPEL, PhaseAffaire.CASSATION, PhaseAffaire.EXECUTION)

        # Legacy locks (keep for backward compat)
        ctx["lock_predecision"] = not has_decision
        ctx["lock_postdecision"] = has_decision

        # Phase indicator data
        phases_ordered = [
            (PhaseAffaire.PRELIMINAIRE, "التمهيدية"),
            (PhaseAffaire.PREMIERE_INSTANCE, "الابتدائية"),
            (PhaseAffaire.APPEL, "الاستئناف"),
            (PhaseAffaire.CASSATION, "النقض"),
            (PhaseAffaire.EXECUTION, "التنفيذ"),
            (PhaseAffaire.CLOTURE, "مقفلة"),
        ]
        current_idx = next((i for i, (val, _) in enumerate(phases_ordered) if val == phase), 0)
        phase_steps = []
        for i, (val, label) in enumerate(phases_ordered):
            if i < current_idx:
                status = "completed"
            elif i == current_idx:
                status = "active"
            else:
                status = "pending"
            phase_steps.append({"value": val, "label": label, "status": status})
        ctx["phase_steps"] = phase_steps
        ctx["current_phase"] = phase

        # Workflow path indicator
        has_avertissement = Avertissement.objects.filter(affaire=affaire).exists()
        if has_avertissement:
            ctx["workflow_path"] = "إنذار → دعوى"
            ctx["workflow_path_code"] = "B"
        elif phase == PhaseAffaire.PRELIMINAIRE:
            ctx["workflow_path"] = "مسار لم يُحدد بعد"
            ctx["workflow_path_code"] = "?"
        else:
            ctx["workflow_path"] = "دعوى مباشرة"
            ctx["workflow_path_code"] = "C"

        # Document checklist
        requirements = DocumentRequirement.objects.filter(phase=phase)
        existing_titles = set(
            PieceJointe.objects.filter(affaire=affaire).values_list("titre", flat=True)
        )
        doc_checklist = []
        for req in requirements:
            doc_checklist.append({
                "nom": req.nom_document,
                "obligatoire": req.obligatoire,
                "present": req.nom_document in existing_titles,
            })
        ctx["doc_checklist"] = doc_checklist

        # Last mahakim sync
        ctx["last_mahakim_sync"] = (
            MahakimSyncResult.objects.filter(affaire=affaire).order_by("-date_sync").first()
        )

        return ctx

class _AffaireFormMixin(HTMXModalFormMixin):
    """Shared logic for AffaireCreate / AffaireUpdate."""
    form_class = AffaireForm
    page_template = "cabinet/affaire_form.html"

    def get_template_names(self):
        if self.htmx():
            return ["modals/_affaire_form.html"]
        return [self.page_template]

    def form_invalid(self, form):
        if self.htmx():
            return self.render_modal("modals/_affaire_form.html", {
                "form": form, "title": "نموذج القضية", "action": self.request.path
            })
        return super().form_invalid(form)

    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json(
                self.success_message,
                redirect=str(reverse_lazy("cabinet:affaire_detail", args=[self.object.pk])),
                closeModal=True,
            )
        messages.success(self.request, self.success_message)
        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy("cabinet:affaire_detail", args=[self.object.pk])

class AffaireCreate(UIPermRequiredMixin, SecureBase, _AffaireFormMixin, CreateView):
    ui_perm = "ui_btn_add"
    model = Affaire
    permission_required = "cabinet.add_affaire"
    success_message = "تم إنشاء القضية."

class AffaireUpdate(UIPermRequiredMixin, SecureBase, _AffaireFormMixin, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Affaire
    permission_required = "cabinet.change_affaire"
    success_message = "تم تحديث القضية."

class AffaireDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Affaire
    permission_required = "cabinet.delete_affaire"
    def get_success_url(self): return reverse_lazy("cabinet:affaire_list")

# =============================================================
# JURIDICTION
# =============================================================

class JuridictionList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = Juridiction
    template_name = "avocat/juridiction_list.html"
    partial_template_name = "avocat/juridiction_table.html"
    permission_required = "cabinet.view_juridiction"
    paginate_by = 20  # optionnel

    # champs cherchables (ar/fr + FK)
    search_fields = [
        "code",
        "nomtribunal_ar", "nomtribunal_fr",
        "villetribunal_ar", "villetribunal_fr",
        "type__libelle", "type__libelle_fr", "type__code_type",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("type", "TribunalParent").order_by("id")


class JuridictionDetail(SecureBase, DetailView):
    model = Juridiction
    template_name = "avocat/juridiction_detail.html"
    permission_required = "cabinet.view_juridiction"

class JuridictionCreate(UIPermRequiredMixin, SecureBase, ModalCreateView, CreateView):
    ui_perm = "ui_btn_add"
    model = Juridiction
    form_class = JuridictionForm
    permission_required = "cabinet.add_juridiction"
    success_message = "تم إنشاء المحكمة."
    page_template = "cabinet/juridiction_form.html"
    def get_success_url(self): return self.request.GET.get("next") or reverse_lazy("cabinet:juridiction_list")

class JuridictionUpdate(UIPermRequiredMixin, SecureBase, ModalUpdateView, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Juridiction
    form_class = JuridictionForm
    permission_required = "cabinet.change_juridiction"
    success_message = "تم تحديث المحكمة."
    page_template = "cabinet/juridiction_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:juridiction_detail", args=[self.object.pk])

class JuridictionDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Juridiction
    permission_required = "cabinet.delete_juridiction"
    def get_success_url(self): return reverse_lazy("cabinet:juridiction_list")

# =============================================================
# AVOCAT & BARREAU
# =============================================================

class AvocatList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = Avocat
    template_name = "avocat/avocat_list.html"
    permission_required = "cabinet.view_avocat"
    paginate_by = 20  # optionnel

    # champs cherchables (ar/fr + FK)
    search_fields = [
        "nom", "barreau"
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("barreau").order_by("id")



class AvocatDetail(SecureBase, DetailView):
    model = Avocat
    template_name = "avocat/avocat_detail.html"
    permission_required = "cabinet.view_avocat"

class AvocatCreate(UIPermRequiredMixin, SecureBase, ModalCreateView, CreateView):
    ui_perm = "ui_btn_add"
    model = Avocat
    form_class = AvocatForm
    permission_required = "cabinet.add_avocat"
    success_message = "تم إضافة المحامي."
    page_template = "avocat/avocat_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:avocat_detail", args=[self.object.pk])

class AvocatUpdate(UIPermRequiredMixin, SecureBase, ModalUpdateView, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Avocat
    form_class = AvocatForm
    permission_required = "cabinet.change_avocat"
    success_message = "تم تعديل بيانات المحامي."
    page_template = "avocat/avocat_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:avocat_detail", args=[self.object.pk])

class AvocatDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Avocat
    permission_required = "cabinet.delete_avocat"
    def get_success_url(self): return reverse_lazy("cabinet:avocat_list")

class BarreauList(SecureBase, SoftDeleteQuerysetMixin, ListView):
    model = Barreau
    template_name = "avocat/barreau_list.html"
    context_object_name = "object_list"
    permission_required = "cabinet.view_barreau"
    paginate_by = 20  # optionnel

    # champs cherchables (ar/fr + FK)
    search_fields = [
        "code",
        "juridiction_appel",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("juridiction_appel").order_by("id")


class BarreauDetail(SecureBase, DetailView):
    model = Barreau
    template_name = "avocat/barreau_detail.html"
    permission_required = "cabinet.view_barreau"

class BarreauCreate(UIPermRequiredMixin, SecureBase, ModalCreateView, CreateView):
    ui_perm = "ui_btn_add"
    model = Barreau
    form_class = BarreauForm
    permission_required = "cabinet.add_barreau"
    success_message = "تمّ حفظ الهيئة."
    page_template = "avocat/barreau_form.html"
    refresh_target = "#ref-list"
    def get_success_url(self): return reverse_lazy("cabinet:barreau_list")

class BarreauUpdate(UIPermRequiredMixin, SecureBase, ModalUpdateView, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Barreau
    form_class = BarreauForm
    permission_required = "cabinet.change_barreau"
    success_message = "تمّ تحديث الهيئة."
    page_template = "avocat/barreau_form.html"
    refresh_target = "#ref-list"
    def get_success_url(self): return reverse_lazy("cabinet:barreau_list")

class BarreauDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Barreau
    permission_required = "cabinet.delete_barreau"
    def get_success_url(self): return reverse_lazy("cabinet:barreau_list")

# =============================================================
# WORKFLOW: Audience / Mesure / Expertise / Decision / Notification
# =============================================================

# ----- Audience
class AudienceList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, DjangoFilterListMixin, HTMXPartialListMixin, ListView):
    model = Audience
    template_name = "avocat/audience_list.html"
    permission_required = "cabinet.view_audience"
    paginate_by = 20
    filterset_class = AudienceFilter

    search_fields = [
        "affaire__reference_interne", "type_audience__libelle",
        "resultat__libelle", "proces_verbal",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("affaire", "type_audience", "resultat").order_by("-date_audience")


class AudienceDetail(SecureBase, DetailView):
    model = Audience
    template_name = "avocat/audience_detail.html"
    permission_required = "cabinet.view_audience"

class AudienceCreate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, CreateView):
    ui_perm = "ui_btn_add"
    model = Audience
    form_class = AudienceForm
    permission_required = "cabinet.add_audience"
    success_message = "تمّ حفظ الجلسة."
    modal_template = "modals/_audience_form.html"
    page_template = "cabinet/audience_form.html"
    def get_initial(self):
        initial = super().get_initial()
        affaire_id = self.request.GET.get("affaire")
        if affaire_id:
            initial["affaire"] = get_object_or_404(Affaire, pk=affaire_id)
        return initial
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ حفظ الجلسة.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[_affaire_pk_from_step(self.object)]),
                                     closeModal=True)
        messages.success(self.request, "تمّ حفظ الجلسة.")
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:audience_list")

class AudienceUpdate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Audience
    form_class = AudienceForm
    permission_required = "cabinet.change_audience"
    success_message = "تمّ تحديث الجلسة."
    modal_template = "modals/_audience_form.html"
    page_template = "cabinet/audience_form.html"
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ تحديث الجلسة.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[_affaire_pk_from_step(self.object)]),
                                     closeModal=True)
        messages.success(self.request, "تمّ تحديث الجلسة.")
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:audience_list")

class AudienceDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Audience
    permission_required = "cabinet.delete_audience"
    def get_success_url(self): return reverse_lazy("cabinet:audience_list")

# ----- Mesure
class MesureList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = Mesure
    template_name = "avocat/mesure_list.html"
    permission_required = "cabinet.view_mesure"
    paginate_by = 20

    search_fields = [
        "type_mesure__libelle", "notes",
        "audience__affaire__reference_interne",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("audience", "type_mesure", "statut").order_by("-date_ordonnee")


class MesureDetail(SecureBase, DetailView):
    model = Mesure
    template_name = "avocat/mesure_detail.html"
    permission_required = "cabinet.view_mesure"

class MesureCreate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, CreateView):
    ui_perm = "ui_btn_add"
    model = Mesure
    form_class = MesureForm
    permission_required = "cabinet.add_mesure"
    success_message = "تمّ حفظ الإجراء."
    page_template = "cabinet/mesure_form.html"
    def get_initial(self):
        initial = super().get_initial()
        aff_pk = self.request.GET.get("affaire")
        if aff_pk:
            aud = (Audience.objects.filter(affaire_id=aff_pk).order_by("-date_audience", "-pk").first())
            if aud:
                initial["audience"] = aud
        return initial
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ حفظ الإجراء.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[_affaire_pk_from_step(self.object)]),
                                     closeModal=True)
        messages.success(self.request, "تمّ حفظ الإجراء.")
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:mesure_list")

class MesureUpdate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Mesure
    form_class = MesureForm
    permission_required = "cabinet.change_mesure"
    success_message = "تمّ تحديث الإجراء."
    page_template = "cabinet/mesure_form.html"
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ تحديث الإجراء.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[_affaire_pk_from_step(self.object)]),
                                     closeModal=True)
        messages.success(self.request, "تمّ تحديث الإجراء.")
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:mesure_list")

class MesureDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Mesure
    permission_required = "cabinet.delete_mesure"
    def get_success_url(self): return reverse_lazy("cabinet:mesure_list")

# ----- Expertise + Expert
class ExpertiseList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = Expertise
    template_name = "avocat/expertise_list.html"
    permission_required = "cabinet.view_expertise"
    paginate_by = 20

    search_fields = [
        "expert_nom", "affaire__reference_interne",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("affaire").order_by("-date_ordonnee")


class ExpertiseDetail(SecureBase, DetailView):
    model = Expertise
    template_name = "avocat/expertise_detail.html"
    permission_required = "cabinet.view_expertise"

class ExpertiseCreate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, CreateView):
    ui_perm = "ui_btn_add"
    model = Expertise
    form_class = ExpertiseForm
    permission_required = "cabinet.add_expertise"
    success_message = "تمّ حفظ الخبرة."
    modal_template = "modals/_expertise_form.html"
    page_template = "cabinet/expertise_form.html"
    def get_initial(self):
        initial = super().get_initial()
        aff = self.request.GET.get("affaire")
        if aff:
            initial["affaire"] = get_object_or_404(Affaire, pk=aff)
        return initial
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ حفظ الخبرة.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[_affaire_pk_from_step(self.object)]),
                                     closeModal=True)
        messages.success(self.request, "تمّ حفظ الخبرة.")
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:expertise_list")

class ExpertiseUpdate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Expertise
    form_class = ExpertiseForm
    permission_required = "cabinet.change_expertise"
    success_message = "تمّ تحديث الخبرة."
    modal_template = "modals/_expertise_form.html"
    page_template = "cabinet/expertise_form.html"
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ تحديث الخبرة.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[_affaire_pk_from_step(self.object)]),
                                     closeModal=True)
        messages.success(self.request, "تمّ تحديث الخبرة.")
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:expertise_list")

class ExpertiseDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Expertise
    permission_required = "cabinet.delete_expertise"
    def get_success_url(self): return reverse_lazy("cabinet:expertise_list")

class ExpertCreate(UIPermRequiredMixin, SecureBase, ModalCreateView, CreateView):
    ui_perm = "ui_btn_add"
    """Création d’un Expert depuis un select2 (+ Ajouter جديد)."""
    model = Expert
    form_class = ExpertForm
    permission_required = "cabinet.add_expert"
    success_message = "تم إضافة الخبير."
    page_template = "cabinet/expert_form.html"
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            select_id = self.request.GET.get("select_id") or self.request.POST.get("select_id")
            # injecter option + sélectionner (compatible select2)
            html = f"""
            <script>
              (function(){{
                 var sel = document.getElementById("{select_id}");
                 if(sel){{
                   var opt = new Option("{self.object.nom}", "{self.object.pk}", true, true);
                   sel.add(opt);
                   if (window.jQuery && jQuery.fn.select2) {{ jQuery(sel).trigger('change'); }}
                 }}
                 var m=document.getElementById('mainModal');
                 if(m) (bootstrap.Modal.getInstance(m)||new bootstrap.Modal(m)).hide();
              }})();
            </script>
            """
            return self.success_json(html=html, closeModal=True)
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:expertise_list")

# ----- Decision
class DecisionList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = Decision
    template_name = "avocat/decision_list.html"
    permission_required = "cabinet.view_decision"
    paginate_by = 20

    search_fields = [
        "numero_decision", "affaire__reference_interne",
        "formation", "resumé",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("affaire").order_by("-date_prononce")


class DecisionDetail(SecureBase, DetailView):
    model = Decision
    template_name = "avocat/decision_detail.html"
    permission_required = "cabinet.view_decision"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            ctx["analysis"] = self.object.analysis
        except DecisionAnalysis.DoesNotExist:
            ctx["analysis"] = None
        ctx["pdf_pieces"] = list(
            self.object.affaire.pieces.filter(type_piece="PDF").order_by("-date_ajout")[:20]
        )
        return ctx

class DecisionCreate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, CreateView):
    ui_perm = "ui_btn_add"
    model = Decision
    form_class = DecisionForm
    permission_required = "cabinet.add_decision"
    success_message = "تمّ حفظ الحكم."
    page_template = "cabinet/decision_form.html"
    def get_initial(self):
        initial = super().get_initial()
        aff = self.request.GET.get("affaire")
        if aff:
            initial["affaire"] = get_object_or_404(Affaire, pk=aff)
        return initial
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ حفظ الحكم.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[_affaire_pk_from_step(self.object)]),
                                     closeModal=True)
        messages.success(self.request, "تمّ حفظ الحكم.")
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:decision_list")

class DecisionUpdate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Decision
    form_class = DecisionForm
    permission_required = "cabinet.change_decision"
    success_message = "تمّ تحديث الحكم."
    page_template = "cabinet/decision_form.html"
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ تحديث الحكم.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[_affaire_pk_from_step(self.object)]),
                                     closeModal=True)
        messages.success(self.request, "تمّ تحديث الحكم.")
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:decision_list")

class DecisionDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Decision
    permission_required = "cabinet.delete_decision"
    def get_success_url(self): return reverse_lazy("cabinet:decision_list")

# ----- Notification
class NotificationList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = Notification
    template_name = "avocat/notification_list.html"
    permission_required = "cabinet.view_notification"
    paginate_by = 20

    search_fields = [
        "demande_numero", "huissier_nom",
        "decision__numero_decision", "decision__affaire__reference_interne",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("decision", "decision__affaire").order_by("-date_depot_demande")


class NotificationDetail(SecureBase, DetailView):
    model = Notification
    template_name = "avocat/notification_detail.html"
    permission_required = "cabinet.view_notification"

class NotificationCreate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, CreateView):
    ui_perm = "ui_btn_add"
    model = Notification
    form_class = NotificationForm
    permission_required = "cabinet.add_notification"
    success_message = "تمّ حفظ التبليغ."
    modal_template = "modals/_notification_form.html"
    page_template = "cabinet/notification_form.html"
    def get_initial(self):
        initial = super().get_initial()
        aff = self.request.GET.get("affaire")
        if aff:
            dec = Decision.objects.filter(affaire_id=aff).order_by("-date_prononce", "-pk").first()
            if dec: initial["decision"] = dec
        return initial
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ حفظ التبليغ.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[_affaire_pk_from_step(self.object)]),
                                     closeModal=True)
        messages.success(self.request, "تمّ حفظ التبليغ.")
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:notification_list")

class NotificationUpdate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Notification
    form_class = NotificationForm
    permission_required = "cabinet.change_notification"
    success_message = "تمّ تحديث التبليغ."
    modal_template = "modals/_notification_form.html"
    page_template = "cabinet/notification_form.html"
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ تحديث التبليغ.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[_affaire_pk_from_step(self.object)]),
                                     closeModal=True)
        messages.success(self.request, "تمّ تحديث التبليغ.")
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:notification_list")

class NotificationDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Notification
    permission_required = "cabinet.delete_notification"
    def get_success_url(self): return reverse_lazy("cabinet:notification_list")

# ----- Voie de recours
class VoieDeRecoursList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = VoieDeRecours
    template_name = "avocat/voiederecours_list.html"
    permission_required = "cabinet.view_voiederecours"
    paginate_by = 20

    search_fields = [
        "numero_recours", "decision__numero_decision",
        "decision__affaire__reference_interne",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("decision", "type_recours", "statut", "juridiction").order_by("-date_depot")


class VoieDeRecoursDetail(SecureBase, DetailView):
    model = VoieDeRecours
    template_name = "avocat/voiederecours_detail.html"
    permission_required = "cabinet.view_voiederecours"

class VoieDeRecoursCreate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, CreateView):
    ui_perm = "ui_btn_add"
    model = VoieDeRecours
    form_class = VoieDeRecoursForm
    permission_required = "cabinet.add_voiederecours"
    success_message = "تمّ حفظ الطعن."
    page_template = "cabinet/voiederecours_form.html"
    def get_initial(self):
        initial = super().get_initial()
        aff = self.request.GET.get("affaire")
        if aff:
            dec = Decision.objects.filter(affaire_id=aff).order_by("-date_prononce", "-pk").first()
            if dec: initial["decision"] = dec
        return initial
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ حفظ الطعن.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[_affaire_pk_from_step(self.object)]),
                                     closeModal=True)
        messages.success(self.request, "تمّ حفظ الطعن.")
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:voiederecours_list")

class VoieDeRecoursUpdate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, UpdateView):
    ui_perm = "ui_btn_edit"
    model = VoieDeRecours
    form_class = VoieDeRecoursForm
    permission_required = "cabinet.change_voiederecours"
    success_message = "تمّ تحديث الطعن."
    page_template = "cabinet/voiederecours_form.html"
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ تحديث الطعن.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[_affaire_pk_from_step(self.object)]),
                                     closeModal=True)
        messages.success(self.request, "تمّ تحديث الطعن.")
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:voiederecours_list")

class VoieDeRecoursDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = VoieDeRecours
    permission_required = "cabinet.delete_voiederecours"
    def get_success_url(self): return reverse_lazy("cabinet:voiederecours_list")

# ----- Exécution
class ExecutionList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = Execution
    template_name = "avocat/execution_list.html"
    permission_required = "cabinet.view_execution"
    paginate_by = 20

    search_fields = [
        "decision__numero_decision", "decision__affaire__reference_interne",
        "type_execution__libelle",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("decision", "type_execution", "statut").order_by("-date_demande")


class ExecutionDetail(SecureBase, DetailView):
    model = Execution
    template_name = "avocat/execution_detail.html"
    permission_required = "cabinet.view_execution"


class ExecutionCreate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, CreateView):
    ui_perm = "ui_btn_add"
    model = Execution
    form_class = ExecutionForm
    permission_required = "cabinet.add_execution"
    success_message = "تمّ حفظ التنفيذ."
    modal_template = "modals/_execution_form.html"
    page_template = "cabinet/execution_form.html"
    def get_initial(self):
        initial = super().get_initial()
        aff = self.request.GET.get("affaire")
        if aff:
            dec = Decision.objects.filter(affaire_id=aff).order_by("-date_prononce", "-pk").first()
            if dec: initial["decision"] = dec
        return initial
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ حفظ التنفيذ.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[_affaire_pk_from_step(self.object)]),
                                     closeModal=True)
        messages.success(self.request, "تمّ حفظ التنفيذ.")
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:execution_list")

class ExecutionUpdate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Execution
    form_class = ExecutionForm
    permission_required = "cabinet.change_execution"
    success_message = "تمّ تحديث التنفيذ."
    modal_template = "modals/_execution_form.html"
    page_template = "cabinet/execution_form.html"
    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ تحديث التنفيذ.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[_affaire_pk_from_step(self.object)]),
                                     closeModal=True)
        messages.success(self.request, "تمّ تحديث التنفيذ.")
        return super().form_valid(form)
    def get_success_url(self): return reverse_lazy("cabinet:execution_list")

class ExecutionDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Execution
    permission_required = "cabinet.delete_execution"
    def get_success_url(self): return reverse_lazy("cabinet:execution_list")

# =============================================================
# AUTRES ENTITÉS SIMPLES
# =============================================================

class AlerteList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = Alerte
    template_name = "avocat/alerte_list.html"
    permission_required = "cabinet.view_alerte"
    paginate_by = 20

    search_fields = [
        "message", "destinataire",
        "type_alerte__libelle",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("type_alerte").order_by("-date_alerte")


class AlerteDetail(SecureBase, DetailView):
    model = Alerte
    template_name = "avocat/alerte_detail.html"
    permission_required = "cabinet.view_alerte"

class AlerteCreate(UIPermRequiredMixin, SecureBase, ModalCreateView, CreateView):
    ui_perm = "ui_btn_add"
    model = Alerte
    form_class = AlerteForm
    permission_required = "cabinet.add_alerte"
    success_message = "تمّ حفظ التنبيه."
    page_template = "cabinet/alerte_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:alerte_list")

class AlerteUpdate(UIPermRequiredMixin, SecureBase, ModalUpdateView, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Alerte
    form_class = AlerteForm
    permission_required = "cabinet.change_alerte"
    success_message = "تمّ تحديث التنبيه."
    page_template = "cabinet/alerte_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:alerte_list")

class AlerteDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Alerte
    permission_required = "cabinet.delete_alerte"
    def get_success_url(self): return reverse_lazy("cabinet:alerte_list")

class TacheList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = Tache
    template_name = "avocat/tache_list.html"
    permission_required = "cabinet.view_tache"
    paginate_by = 20

    search_fields = [
        "titre", "description",
        "affaire__reference_interne", "assigne_a__nom_complet",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("affaire", "assigne_a", "statut").order_by("-echeance")


class TacheDetail(SecureBase, DetailView):
    model = Tache
    template_name = "avocat/tache_detail.html"
    permission_required = "cabinet.view_tache"

class TacheCreate(UIPermRequiredMixin, SecureBase, ModalCreateView, CreateView):
    ui_perm = "ui_btn_add"
    model = Tache
    form_class = TacheForm
    permission_required = "cabinet.add_tache"
    success_message = "تمّ حفظ المهمة."
    page_template = "cabinet/tache_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:tache_list")

class TacheUpdate(UIPermRequiredMixin, SecureBase, ModalUpdateView, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Tache
    form_class = TacheForm
    permission_required = "cabinet.change_tache"
    success_message = "تمّ تحديث المهمة."
    page_template = "cabinet/tache_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:tache_list")

class TacheDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Tache
    permission_required = "cabinet.delete_tache"
    def get_success_url(self): return reverse_lazy("cabinet:tache_list")

class PieceJointeList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, DjangoFilterListMixin, HTMXPartialListMixin, ListView):
    model = PieceJointe
    template_name = "avocat/piecejointe_list.html"
    permission_required = "cabinet.view_piecejointe"
    paginate_by = 20
    filterset_class = PieceJointeFilter

    search_fields = [
        "titre", "affaire__reference_interne",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("affaire").order_by("-date_ajout")


class PieceJointeDetail(SecureBase, DetailView):
    model = PieceJointe
    template_name = "avocat/piecejointe_detail.html"
    permission_required = "cabinet.view_piecejointe"

class PieceJointeCreate(UIPermRequiredMixin, SecureBase, ModalCreateView, CreateView):
    ui_perm = "ui_btn_add"
    model = PieceJointe
    form_class = PieceJointeForm
    permission_required = "cabinet.add_piecejointe"
    success_message = "تمّ حفظ المرفق."
    page_template = "cabinet/piecejointe_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:piecejointe_list")

class PieceJointeUpdate(UIPermRequiredMixin, SecureBase, ModalUpdateView, UpdateView):
    ui_perm = "ui_btn_edit"
    model = PieceJointe
    form_class = PieceJointeForm
    permission_required = "cabinet.change_piecejointe"
    success_message = "تمّ تحديث المرفق."
    page_template = "cabinet/piecejointe_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:piecejointe_list")

class PieceJointeDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = PieceJointe
    permission_required = "cabinet.delete_piecejointe"
    def get_success_url(self): return reverse_lazy("cabinet:piecejointe_list")

# ----- Parties / Liens Affaire<->Partie/Avocat
class PartieList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = Partie
    template_name = "avocat/partie_list.html"
    permission_required = "cabinet.view_partie"
    paginate_by = 20

    search_fields = [
        "nom_complet", "cin_ou_rc", "telephone", "email",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("avocat").order_by("nom_complet")


class PartieDetail(SecureBase, DetailView):
    model = Partie
    template_name = "avocat/partie_detail.html"
    permission_required = "cabinet.view_partie"

class PartieCreate(UIPermRequiredMixin, SecureBase, ModalCreateView, CreateView):
    ui_perm = "ui_btn_add"
    model = Partie
    form_class = PartieForm
    permission_required = "cabinet.add_partie"
    success_message = "تمّ حفظ الطرف."
    page_template = "cabinet/partie_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:partie_list")

class PartieUpdate(UIPermRequiredMixin, SecureBase, ModalUpdateView, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Partie
    form_class = PartieForm
    permission_required = "cabinet.change_partie"
    success_message = "تمّ تحديث الطرف."
    page_template = "cabinet/partie_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:partie_list")

class PartieDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Partie
    permission_required = "cabinet.delete_partie"
    def get_success_url(self): return reverse_lazy("cabinet:partie_list")

class AffaireAvocatList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = AffaireAvocat
    template_name = "avocat/affaireavocat_list.html"
    permission_required = "cabinet.view_affaireavocat"
    paginate_by = 20

    search_fields = [
        "avocat__nom", "affaire__reference_interne",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("affaire", "avocat").order_by("id")


class AffaireAvocatDetail(SecureBase, DetailView):
    model = AffaireAvocat
    template_name = "avocat/affaireavocat_detail.html"
    permission_required = "cabinet.view_affaireavocat"

class AffaireAvocatCreate(UIPermRequiredMixin, SecureBase, ModalCreateView, CreateView):
    ui_perm = "ui_btn_add"
    model = AffaireAvocat
    form_class = AffaireAvocatForm
    permission_required = "cabinet.add_affaireavocat"
    success_message = "تمّ ربط المحامي بالملف."
    page_template = "cabinet/affaireavocat_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:affaireavocat_list")

class AffaireAvocatUpdate(UIPermRequiredMixin, SecureBase, ModalUpdateView, UpdateView):
    ui_perm = "ui_btn_edit"
    model = AffaireAvocat
    form_class = AffaireAvocatForm
    permission_required = "cabinet.change_affaireavocat"
    success_message = "تمّ تحديث الربط."
    page_template = "cabinet/affaireavocat_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:affaireavocat_list")

class AffaireAvocatDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = AffaireAvocat
    permission_required = "cabinet.delete_affaireavocat"
    def get_success_url(self): return reverse_lazy("cabinet:affaireavocat_list")

class AffairePartieList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = AffairePartie
    template_name = "avocat/affairepartie_list.html"
    permission_required = "cabinet.view_affairepartie"
    paginate_by = 20

    search_fields = [
        "partie__nom_complet", "affaire__reference_interne",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("affaire", "partie").order_by("id")


class AffairePartieDetail(SecureBase, DetailView):
    model = AffairePartie
    template_name = "avocat/affairepartie_detail.html"
    permission_required = "cabinet.view_affairepartie"

class AffairePartieCreate(UIPermRequiredMixin, SecureBase, ModalCreateView, CreateView):
    ui_perm = "ui_btn_add"
    model = AffairePartie
    form_class = AffairePartieForm
    permission_required = "cabinet.add_affairepartie"
    success_message = "تمّ ربط الطرف بالملف."
    page_template = "cabinet/affairepartie_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:affairepartie_list")

class AffairePartieUpdate(UIPermRequiredMixin, SecureBase, ModalUpdateView, UpdateView):
    ui_perm = "ui_btn_edit"
    model = AffairePartie
    form_class = AffairePartieForm
    permission_required = "cabinet.change_affairepartie"
    success_message = "تمّ تحديث الربط."
    page_template = "cabinet/affairepartie_form.html"
    def get_success_url(self): return reverse_lazy("cabinet:affairepartie_list")

class AffairePartieDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = AffairePartie
    permission_required = "cabinet.delete_affairepartie"
    def get_success_url(self): return reverse_lazy("cabinet:affairepartie_list")

# =============================================================
# PARTIAL: TIMELINE D’UNE AFFAIRE (HTMX)
# =============================================================

def affaire_timeline_partial(request, pk):
    affaire = get_object_or_404(Affaire, pk=pk)
    events: list[dict] = []
    today = timezone.localdate()
    _vb = reverse("cabinet:document_viewer")

    # إنذارات
    for av in Avertissement.objects.filter(affaire=affaire).select_related("type_avertissement").order_by("date_envoi"):
        deadline_info = ""
        deadline_badge = ""
        if av.date_echeance:
            j = av.jours_restants
            if j is not None:
                if j <= 0:
                    deadline_info = "انتهى الأجل"
                    deadline_badge = "danger"
                elif j <= 5:
                    deadline_info = f"باقي {j} أيام"
                    deadline_badge = "danger"
                elif j <= 10:
                    deadline_info = f"باقي {j} يوم"
                    deadline_badge = "warning"
                else:
                    deadline_info = f"باقي {j} يوم"
                    deadline_badge = "success"
        _av_file = bool(av.document) or bool(av.preuve_envoi)
        _av_url = ""
        if av.document:
            _av_url = f"{_vb}?model=avertissement&pk={av.pk}&field=document"
        elif av.preuve_envoi:
            _av_url = f"{_vb}?model=avertissement&pk={av.pk}&field=preuve_envoi"
        events.append({
            "date": av.date_envoi, "date_str": ar_dt(av.date_envoi),
            "type": "إنذار", "title": str(av.type_avertissement),
            "meta": f"{av.destinataire_nom} — {av.get_resultat_display()}",
            "badge": "terracotta",
            "deadline_info": deadline_info, "deadline_badge": deadline_badge,
            "has_file": _av_file, "viewer_url": _av_url,
        })

    # جلسات
    for a in Audience.objects.filter(affaire=affaire).select_related("type_audience", "resultat").order_by("date_audience"):
        events.append({
            "date": a.date_audience, "date_str": ar_dt(a.date_audience),
            "type": "جلسة", "title": str(a.type_audience),
            "meta": str(a.resultat) if a.resultat_id else "", "badge": "warning",
            "has_file": False, "viewer_url": "",
        })
    # إجراءات
    for m in Mesure.objects.filter(audience__affaire=affaire).select_related("type_mesure", "statut").order_by("id"):
        dt = getattr(m, "date_ordonnee", None) or getattr(m, "date", None)
        events.append({
            "date": dt, "date_str": ar_dt(dt),
            "type": "إجراء",
            "title": str(m.type_mesure),
            "meta": str(m.statut) if m.statut_id else "",
            "badge": "primary",
            "has_file": False, "viewer_url": "",
        })
    # خبرات
    for e in Expertise.objects.filter(affaire=affaire).order_by("date_ordonnee", "date_depot"):
        _exp_file = bool(e.rapport)
        _exp_url = f"{_vb}?model=expertise&pk={e.pk}&field=rapport" if _exp_file else ""
        if e.date_ordonnee:
            events.append({"date": e.date_ordonnee, "date_str": ar_dt(e.date_ordonnee),
                           "type": "خبرة", "title": f"أمرت: {e.expert_nom or ''}",
                           "meta": "خبرة مضادة" if getattr(e, "contre_expertise", False) else "", "badge": "dark",
                           "has_file": _exp_file, "viewer_url": _exp_url})
        if e.date_depot:
            events.append({"date": e.date_depot, "date_str": ar_dt(e.date_depot),
                           "type": "خبرة", "title": "إيداع الخبرة",
                           "meta": e.expert_nom or "", "badge": "dark",
                           "has_file": _exp_file, "viewer_url": _exp_url})
    # أحكام
    for d in Decision.objects.filter(affaire=affaire).order_by("date_prononce"):
        events.append({"date": d.date_prononce, "date_str": ar_dt(d.date_prononce),
                       "type": "حكم", "title": f"حكم رقم {d.numero_decision or ''}",
                       "meta": "قابل للطعن" if getattr(d, "susceptible_recours", False) else "غير قابل للطعن",
                       "badge": "secondary", "is_phase_marker": True,
                       "has_file": False, "viewer_url": ""})
    # تبليغات
    for n in Notification.objects.filter(decision__affaire=affaire).order_by("date_signification"):
        _not_file = bool(n.preuve)
        _not_url = f"{_vb}?model=notification&pk={n.pk}&field=preuve" if _not_file else ""
        events.append({"date": n.date_signification, "date_str": ar_dt(n.date_signification),
                       "type": "تبليغ", "title": f"طلب {n.demande_numero or ''}",
                       "meta": f"مفوض: {n.huissier_nom or ''}", "badge": "info",
                       "has_file": _not_file, "viewer_url": _not_url})
    # طرق الطعن
    for r in VoieDeRecours.objects.filter(decision__affaire=affaire).select_related("type_recours", "statut").order_by("date_depot"):
        deadline_info = ""
        deadline_badge = ""
        if r.date_echeance_recours:
            j = r.jours_restants_recours
            if j is not None:
                if j <= 0:
                    deadline_info = "انتهى الأجل"
                    deadline_badge = "danger"
                elif j <= 5:
                    deadline_info = f"باقي {j} أيام"
                    deadline_badge = "danger"
                elif j <= 10:
                    deadline_info = f"باقي {j} يوم"
                    deadline_badge = "warning"
                else:
                    deadline_info = f"باقي {j} يوم"
                    deadline_badge = "success"
        events.append({"date": r.date_depot, "date_str": ar_dt(r.date_depot),
                       "type": "طعن", "title": str(r.type_recours),
                       "meta": str(r.statut) if r.statut_id else "", "badge": "success",
                       "deadline_info": deadline_info, "deadline_badge": deadline_badge,
                       "has_file": False, "viewer_url": ""})
    # تنفيذ
    for ex in Execution.objects.filter(decision__affaire=affaire).select_related("type_execution", "statut").order_by("date_demande"):
        events.append({"date": ex.date_demande, "date_str": ar_dt(ex.date_demande),
                       "type": "تنفيذ", "title": str(ex.type_execution),
                       "meta": str(ex.statut) if ex.statut_id else "", "badge": "success",
                       "has_file": False, "viewer_url": ""})
    # مرفقات
    for p in PieceJointe.objects.filter(affaire=affaire).order_by("date_ajout"):
        _pj_file = bool(p.fichier)
        _pj_url = f"{_vb}?model=piecejointe&pk={p.pk}" if _pj_file else ""
        events.append({"date": p.date_ajout, "date_str": ar_dt(p.date_ajout),
                       "type": "مرفق", "title": p.titre or "",
                       "meta": getattr(p, "get_type_piece_display", lambda: (p.type_piece or ""))(), "badge": "light",
                       "has_file": _pj_file, "viewer_url": _pj_url})

    from datetime import datetime as _dt, date as _d
    def _sort_key(ev):
        d = ev["date"]
        if isinstance(d, _d) and not isinstance(d, _dt):
            return _dt(d.year, d.month, d.day)
        return d.replace(tzinfo=None) if d.tzinfo else d

    events = [ev for ev in events if ev["date"]]
    events.sort(key=_sort_key)
    return render(request, "affaires/_timeline.html", {"affaire": affaire, "events": events})

class ExpertList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = Expert
    template_name = 'avocat/expert_list.html'
    permission_required = 'cabinet.view_expert'
    paginate_by = 20

    search_fields = [
        "nom_complet", "specialite", "email",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.order_by("nom_complet")


class ExpertDetail(SecureBase, DetailView):
    model = Expert
    template_name = 'avocat/expert_detail.html'
    permission_required = 'cabinet.view_expert'
class ExpertUpdate(UIPermRequiredMixin, SecureBase, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Expert
    form_class = ExpertForm
    permission_required = 'cabinet.change_expert'
    success_url = reverse_lazy('cabinet:expert_list')

    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, 'تمّ تحديث الخبرة.')
        if self.htmx():
            return self.success_json('تمّ تحديث الخبرة.', _affaire_pk_from_step(self.object))
        return redirect(self.success_url)

    def get(self, request, *args, **kwargs):
        if self.htmx():
            self.object = self.get_object()
            form = self.form_class(instance=self.object)
            return self.render_modal('modals/_form.html', {'form': form, 'title': 'تعديل خبرة', 'action': request.path})
        return super().get(request, *args, **kwargs)

class ExpertDelete(UIPermRequiredMixin, SecureBase, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Expert
    permission_required = 'cabinet.delete_expert'
    success_url = reverse_lazy('cabinet:expert_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        affaire_pk = _affaire_pk_from_step(self.object)
        self.object.delete()
        messages.success(request, 'تمّ حذف الخبرة.')
        if self.htmx():
            return self.success_json('تمّ حذف الخبرة.', affaire_pk)
        return redirect(self.success_url)

    def get(self, request, *args, **kwargs):
        if self.htmx():
            self.object = self.get_object()
            return self.render_modal('modals/_confirm.html', {'title': 'تأكيد الحذف', 'action': request.path})
        return super().get(request, *args, **kwargs)

class UtilisateurList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = Utilisateur
    template_name = 'avocat/utilisateur_list.html'
    permission_required = 'cabinet.view_utilisateur'
    paginate_by = 20

    search_fields = [
        "nom_complet",
        "email",
        "role__libelle", "role__libelle_fr",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("role").order_by("nom_complet")


class UtilisateurDetail(SecureBase, DetailView):
    model = Utilisateur
    template_name = 'avocat/utilisateur_detail.html'
    permission_required = 'cabinet.view_utilisateur'

class UtilisateurCreate(UIPermRequiredMixin, SecureBase, CreateView):
    ui_perm = "ui_btn_add"
    model = Utilisateur
    form_class = UtilisateurForm
    permission_required = 'cabinet.add_utilisateur'
    success_url = reverse_lazy('cabinet:utilisateur_list')

    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, 'تمّ إضافة المستخدم.')
        if self.htmx():
            return self.success_json('تمّ إضافة المستخدم.')
        return redirect(self.success_url)

    def get(self, request, *args, **kwargs):
        if self.htmx():
            form = self.form_class()
            return self.render_modal('modals/_form.html', {'form': form, 'title': 'إضافة مستخدم', 'action': request.path})
        return super().get(request, *args, **kwargs)

class UtilisateurUpdate(UIPermRequiredMixin, SecureBase, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Utilisateur
    form_class = UtilisateurForm
    permission_required = 'cabinet.change_utilisateur'
    success_url = reverse_lazy('cabinet:utilisateur_list')

    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, 'تمّ تحديث المستخدم.')
        if self.htmx():
            return self.success_json('تمّ تحديث المستخدم.')
        return redirect(self.success_url)

    def get(self, request, *args, **kwargs):
        if self.htmx():
            self.object = self.get_object()
            form = self.form_class(instance=self.object)
            return self.render_modal('modals/_form.html', {'form': form, 'title': 'تعديل مستخدم', 'action': request.path})
        return super().get(request, *args, **kwargs)

class UtilisateurDelete(UIPermRequiredMixin, SecureBase, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Utilisateur
    permission_required = 'cabinet.delete_utilisateur'
    success_url = reverse_lazy('cabinet:utilisateur_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.delete()
        messages.success(request, 'تمّ حذف المستخدم.')
        if self.htmx():
            return self.success_json('تمّ حذف المستخدم.')
        return redirect(self.success_url)

    def get(self, request, *args, **kwargs):
        if self.htmx():
            self.object = self.get_object()
            return self.render_modal('modals/_confirm.html', {'title': 'تأكيد الحذف', 'action': request.path})
        return super().get(request, *args, **kwargs)


class RecetteList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, DjangoFilterListMixin, HTMXPartialListMixin, ListView):
    model = Recette
    template_name = 'avocat/recette_list.html'
    permission_required = 'cabinet.view_recette'
    paginate_by = 20
    filterset_class = RecetteFilter

    search_fields = [
        "affaire__reference_interne",
        "type_recette__libelle", "type_recette__libelle_fr",
        "source",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("affaire", "type_recette").order_by("-date_recette")


class RecetteDetail(SecureBase, DetailView):
    model = Recette
    template_name = 'avocat/recette_detail.html'
    permission_required = 'cabinet.view_recette'

class RecetteCreate(UIPermRequiredMixin, SecureBase, CreateView):
    ui_perm = "ui_btn_add"
    model = Recette
    form_class = RecetteForm
    permission_required = 'cabinet.add_recette'
    success_url = reverse_lazy('cabinet:recette_list')

    def get_initial(self):
        initial = super().get_initial()
        affaire_id = self.request.GET.get('affaire')
        if affaire_id:
            initial['affaire'] = affaire_id
        return initial

    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, 'تمّ حفظ الدخل.')
        if self.htmx():
            return self.success_json('تمّ حفظ الدخل.', _affaire_pk_from_step(self.object))
        return redirect(self.success_url)

    def get(self, request, *args, **kwargs):
        if self.htmx():
            form = self.form_class(initial=self.get_initial())
            return self.render_modal('modals/_form.html', {'form': form, 'title': 'إضافة دخل', 'action': request.path})
        return super().get(request, *args, **kwargs)

class RecetteUpdate(UIPermRequiredMixin, SecureBase, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Recette
    form_class = RecetteForm
    permission_required = 'cabinet.change_recette'
    success_url = reverse_lazy('cabinet:recette_list')

    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, 'تمّ تحديث الدخل.')
        if self.htmx():
            return self.success_json('تمّ تحديث الدخل.', _affaire_pk_from_step(self.object))
        return redirect(self.success_url)

    def get(self, request, *args, **kwargs):
        if self.htmx():
            self.object = self.get_object()
            form = self.form_class(instance=self.object)
            return self.render_modal('modals/_form.html', {'form': form, 'title': 'تعديل دخل', 'action': request.path})
        return super().get(request, *args, **kwargs)

class RecetteDelete(UIPermRequiredMixin, SecureBase, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Recette
    permission_required = 'cabinet.delete_recette'
    success_url = reverse_lazy('cabinet:recette_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        affaire_pk = _affaire_pk_from_step(self.object)
        self.object.delete()
        messages.success(request, 'تمّ حذف الدخل.')
        if self.htmx():
            return self.success_json('تمّ حذف الدخل.', affaire_pk)
        return redirect(self.success_url)

    def get(self, request, *args, **kwargs):
        if self.htmx():
            self.object = self.get_object()
            return self.render_modal('modals/_confirm.html', {'title': 'تأكيد الحذف', 'action': request.path})
        return super().get(request, *args, **kwargs)


class DepenseList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, DjangoFilterListMixin, HTMXPartialListMixin, ListView):
    model = Depense
    template_name = 'avocat/depense_list.html'
    permission_required = 'cabinet.view_depense'
    paginate_by = 20
    filterset_class = DepenseFilter

    search_fields = [
        "affaire__reference_interne",
        "type_depense__libelle", "type_depense__libelle_fr",
        "beneficiaire",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("affaire", "type_depense").order_by("-date_depense")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Stats on the FILTERED queryset (respects active filters)
        fs = getattr(self, "filterset", None)
        stats_qs = fs.qs if fs else self.get_queryset()

        # Total
        agg = stats_qs.aggregate(total=Sum("montant"))
        ctx["depense_total"] = agg["total"] or 0

        # By type
        by_type = (
            stats_qs.values("type_depense__libelle")
            .annotate(total=Sum("montant"))
            .order_by("-total")[:10]
        )
        ctx["stats_by_type"] = [
            {"label": r["type_depense__libelle"] or "—", "total": r["total"]}
            for r in by_type
        ]

        # By beneficiary
        by_benef = (
            stats_qs.exclude(beneficiaire__isnull=True).exclude(beneficiaire="")
            .values("beneficiaire")
            .annotate(total=Sum("montant"))
            .order_by("-total")[:10]
        )
        ctx["stats_by_beneficiaire"] = [
            {"label": r["beneficiaire"], "total": r["total"]}
            for r in by_benef
        ]

        return ctx


class DepenseDetail(SecureBase, DetailView):
    model = Depense
    template_name = 'avocat/depense_detail.html'
    permission_required = 'cabinet.view_depense'

class DepenseCreate(UIPermRequiredMixin, SecureBase, CreateView):
    ui_perm = "ui_btn_add"
    model = Depense
    form_class = DepenseForm
    permission_required = 'cabinet.add_depense'
    success_url = reverse_lazy('cabinet:depense_list')

    def get_initial(self):
        initial = super().get_initial()
        affaire_id = self.request.GET.get('affaire')
        if affaire_id:
            initial['affaire'] = affaire_id
        return initial

    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, 'تمّ حفظ المصروف.')
        if self.htmx():
            return self.success_json('تمّ حفظ المصروف.', _affaire_pk_from_step(self.object))
        return redirect(self.success_url)

    def get(self, request, *args, **kwargs):
        if self.htmx():
            form = self.form_class(initial=self.get_initial())
            return self.render_modal('modals/_form.html', {'form': form, 'title': 'إضافة مصروف', 'action': request.path})
        return super().get(request, *args, **kwargs)

class DepenseUpdate(UIPermRequiredMixin, SecureBase, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Depense
    form_class = DepenseForm
    permission_required = 'cabinet.change_depense'
    success_url = reverse_lazy('cabinet:depense_list')

    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, 'تمّ تحديث المصروف.')
        if self.htmx():
            return self.success_json('تمّ تحديث المصروف.', _affaire_pk_from_step(self.object))
        return redirect(self.success_url)

    def get(self, request, *args, **kwargs):
        if self.htmx():
            self.object = self.get_object()
            form = self.form_class(instance=self.object)
            return self.render_modal('modals/_form.html', {'form': form, 'title': 'تعديل مصروف', 'action': request.path})
        return super().get(request, *args, **kwargs)

class DepenseDelete(UIPermRequiredMixin, SecureBase, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Depense
    permission_required = 'cabinet.delete_depense'
    success_url = reverse_lazy('cabinet:depense_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        affaire_pk = _affaire_pk_from_step(self.object)
        self.object.delete()
        messages.success(request, 'تمّ حذف المصروف.')
        if self.htmx():
            return self.success_json('تمّ حذف المصروف.', affaire_pk)
        return redirect(self.success_url)

    def get(self, request, *args, **kwargs):
        if self.htmx():
            self.object = self.get_object()
            return self.render_modal('modals/_confirm.html', {'title': 'تأكيد الحذف', 'action': request.path})
        return super().get(request, *args, **kwargs)



# -------------------------------------------------------------
# Partial HTMX pour Timeline (déjà référencé par urls.py)
# -------------------------------------------------------------
def ar_dt(dt):
    if not dt:
        return ""
    from datetime import datetime as _dt
    fmt = "DATETIME_FORMAT" if isinstance(dt, _dt) else "DATE_FORMAT"
    return date_format(dt, fmt, use_l10n=True)

#-- Créations imbriquées par Affaire ----
class AudienceCreateForAffaire(SecureBase, HTMXModalFormMixin, CreateView):
    model = Audience            # جلسة
    form_class = AudienceForm
    modal_template = "modals/_audience_form.html"
    template_name = "cabinet/audience_form.html"
    permission_required = "cabinet.add_audience"
    success_message = "تمت إضافة الجلسة."

    def get_affaire(self):
        affaire_pk = self.kwargs.get("affaire_pk")
        return get_object_or_404(Affaire, pk=affaire_pk)

    def dispatch(self, request, *args, **kwargs):
        self.affaire = self.get_affaire()
        if self.affaire.has_decision():  # verrou post-décision
            return self.handle_no_permission()
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.affaire = self.get_affaire()  # ← injecte l'affaire
        obj.save()
        self.object = obj
        return self.success_json("تمت إضافة الجلسة.", refreshTarget="#timeline",
                                 refreshUrl=reverse_lazy("cabinet:affaire_timeline", args=[self.kwargs["affaire_id"]]))

class MesureCreateForAffaire(SecureBase, HTMXModalFormMixin, CreateView):
    model = Mesure              # إجراء
    form_class = MesureForm
    template_name = "cabinet/mesure_form.html"
    permission_required = "cabinet.add_mesure"
    success_message = "تمت إضافة الإجراء."

class ExpertiseCreateForAffaire(SecureBase, HTMXModalFormMixin, CreateView):
    model = Expertise           # خبرة
    form_class = ExpertiseForm
    modal_template = "modals/_expertise_form.html"
    template_name = "cabinet/expertise_form.html"
    permission_required = "cabinet.add_expertise"
    success_message = "تمت إضافة الخبرة."

class DecisionCreateForAffaire(SecureBase, HTMXModalFormMixin, CreateView):
    model = Decision            # حكم
    form_class = DecisionForm
    template_name = "cabinet/decision_form.html"
    permission_required = "cabinet.add_decision"
    success_message = "تمت إضافة الحكم."

class NotificationCreateForAffaire(SecureBase, HTMXModalFormMixin, CreateView):
    model = Notification        # تبليغ
    form_class = NotificationForm
    modal_template = "modals/_notification_form.html"
    template_name = "cabinet/notification_form.html"
    permission_required = "cabinet.add_notification"
    success_message = "تمت إضافة التبليغ."

class RecoursCreateForAffaire(SecureBase, HTMXModalFormMixin, CreateView):
    model = VoieDeRecours       # طعن
    form_class = VoieDeRecoursForm
    template_name = "cabinet/voiederecours_form.html"
    permission_required = "cabinet.add_voiederecours"
    success_message = "تم تسجيل الطعن."

class ExecutionCreateForAffaire(SecureBase, HTMXModalFormMixin, CreateView):
    model = Execution           # تنفيذ
    form_class = ExecutionForm
    modal_template = "modals/_execution_form.html"
    template_name = "cabinet/execution_form.html"
    permission_required = "cabinet.add_execution"
    success_message = "تم فتح ملف التنفيذ."

# =============================================================
# AVERTISSEMENT (إنذار)
# =============================================================

class AvertissementList(SecureBase, NoPostOnReadOnlyMixin, SearchListMixin, HTMXPartialListMixin, ListView):
    model = Avertissement
    template_name = "avocat/avertissement_list.html"
    permission_required = "cabinet.view_avertissement"
    paginate_by = 20

    search_fields = [
        "affaire__reference_interne", "type_avertissement__libelle",
        "destinataire_nom", "objet_avertissement",
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related("affaire", "type_avertissement").order_by("-date_envoi")


class AvertissementDetail(SecureBase, DetailView):
    model = Avertissement
    template_name = "avocat/avertissement_detail.html"
    permission_required = "cabinet.view_avertissement"


class AvertissementCreate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, CreateView):
    ui_perm = "ui_btn_add"
    model = Avertissement
    form_class = AvertissementForm
    permission_required = "cabinet.add_avertissement"
    success_message = "تمّ حفظ الإنذار."
    modal_template = "modals/_avertissement_form.html"
    page_template = "cabinet/avertissement_form.html"

    def get_initial(self):
        initial = super().get_initial()
        aff = self.request.GET.get("affaire")
        if aff:
            initial["affaire"] = get_object_or_404(Affaire, pk=aff)
        return initial

    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ حفظ الإنذار.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[self.object.affaire_id]),
                                     closeModal=True)
        messages.success(self.request, "تمّ حفظ الإنذار.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("cabinet:avertissement_list")


class AvertissementUpdate(UIPermRequiredMixin, SecureBase, HTMXModalFormMixin, UpdateView):
    ui_perm = "ui_btn_edit"
    model = Avertissement
    form_class = AvertissementForm
    permission_required = "cabinet.change_avertissement"
    success_message = "تمّ تحديث الإنذار."
    modal_template = "modals/_avertissement_form.html"
    page_template = "cabinet/avertissement_form.html"

    def form_valid(self, form):
        self.object = form.save()
        if self.htmx():
            return self.success_json("تمّ تحديث الإنذار.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[self.object.affaire_id]),
                                     closeModal=True)
        messages.success(self.request, "تمّ تحديث الإنذار.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("cabinet:avertissement_list")


class AvertissementDelete(UIPermRequiredMixin, SecureBase, ModalDeleteView, DeleteView):
    ui_perm = "ui_btn_delete"
    model = Avertissement
    permission_required = "cabinet.delete_avertissement"
    def get_success_url(self): return reverse_lazy("cabinet:avertissement_list")


class AvertissementCreateForAffaire(SecureBase, HTMXModalFormMixin, CreateView):
    model = Avertissement
    form_class = AvertissementForm
    modal_template = "modals/_avertissement_form.html"
    template_name = "cabinet/avertissement_form.html"
    permission_required = "cabinet.add_avertissement"
    success_message = "تمت إضافة الإنذار."

    def get_affaire(self):
        return get_object_or_404(Affaire, pk=self.kwargs.get("affaire_id"))

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.affaire = self.get_affaire()
        obj.save()
        self.object = obj
        if self.htmx():
            return self.success_json("تمت إضافة الإنذار.", refreshTarget="#timeline",
                                     refreshUrl=reverse("cabinet:affaire_timeline", args=[self.kwargs["affaire_id"]]),
                                     closeModal=True)
        messages.success(self.request, "تمت إضافة الإنذار.")
        return redirect(reverse("cabinet:affaire_detail", args=[self.kwargs["affaire_id"]]))

    def get_success_url(self):
        return reverse_lazy("cabinet:affaire_detail", args=[self.kwargs["affaire_id"]])


# =============================================================
# MAHAKIM.MA SYNC — مزامنة مع بوابة محاكم
# =============================================================

import threading
import logging as _logging
import uuid as _uuid
from io import BytesIO

_mahakim_logger = _logging.getLogger(__name__)

# --- Background task tracking for fetch_mahakim_ids ---
_mahakim_fetch_lock = threading.Lock()
_mahakim_fetch_tasks = {}  # task_id → {status, phase, current, total, name, message, result, errors}


@login_required
def mahakim_preview_single(request, pk):
    """Preview modal before syncing a single affaire."""
    affaire = get_object_or_404(Affaire, pk=pk)
    last_sync = MahakimSyncResult.objects.filter(affaire=affaire).order_by("-date_sync").first()
    html = render_to_string("modals/_mahakim_preview.html", {
        "mode": "preview_single",
        "affaire": affaire,
        "last_sync": last_sync,
    }, request=request)
    return HttpResponse(html)


@login_required
def mahakim_preview_all(request):
    """Preview modal before syncing all affaires."""
    qs = Affaire.objects.filter(
        numero_dossier__isnull=False, code_categorie__isnull=False, annee_dossier__isnull=False,
    ).exclude(numero_dossier="").exclude(annee_dossier="")
    total = qs.count()
    synced_ids = set(MahakimSyncResult.objects.filter(
        affaire__in=qs, success=True
    ).values_list("affaire_id", flat=True).distinct())
    html = render_to_string("modals/_mahakim_preview.html", {
        "mode": "preview_all",
        "total_syncable": total,
        "already_synced": len(synced_ids),
        "never_synced": total - len(synced_ids),
    }, request=request)
    return HttpResponse(html)


def sync_affaire_mahakim(request, pk):
    """Sync une seule affaire avec mahakim.ma (HTMX button)."""
    if not request.user.is_authenticated:
        return JsonResponse({"ok": False, "message": "غير مسموح"}, status=403)

    affaire = get_object_or_404(Affaire, pk=pk)

    if not all([affaire.numero_dossier, affaire.code_categorie, affaire.annee_dossier]):
        return JsonResponse({
            "ok": False,
            "message": "يجب ملء رقم الملف وصنف القضية والسنة قبل المزامنة مع محاكم"
        })

    try:
        from .services.mahakim_scraper import MahakimScraper

        id_mahakim = None
        is_premiere = False
        nom_tribunal = None
        nom_tribunal_appel = None
        if affaire.juridiction:
            id_mahakim = affaire.juridiction.id_mahakim
            is_premiere = bool(affaire.juridiction.TribunalParent)
            nom_tribunal = affaire.juridiction.nomtribunal_ar
            if affaire.juridiction.TribunalParent:
                nom_tribunal_appel = affaire.juridiction.TribunalParent.nomtribunal_ar

        with MahakimScraper(headless=False, timeout=30) as scraper:
            result = scraper.scrape_affaire(
                numero=affaire.numero_dossier,
                code_categorie=affaire.code_categorie.code,
                annee=affaire.annee_dossier,
                id_mahakim_tribunal=id_mahakim,
                is_premiere_instance=is_premiere,
                nom_tribunal=nom_tribunal,
                nom_tribunal_appel=nom_tribunal_appel,
            )


        sync_obj = MahakimSyncResult.objects.create(
            affaire=affaire,
            sync_type='dossier',
            statut_mahakim=result.get("statut_mahakim"),
            prochaine_audience=result.get("prochaine_audience"),
            juge=result.get("juge"),
            observations=result.get("observations"),
            raw_html=(result.get("raw_html") or "")[:50000],
            success=result.get("success", False),
            error_message=result.get("error_message"),
            procedures_json=result.get("procedures") or None,
            parties_json=result.get("parties") or None,
        )

        if result["success"]:
            return JsonResponse({
                "ok": True,
                "message": f"تمت المزامنة بنجاح — الحالة: {result.get('statut_mahakim', '—')}",
                "data": {
                    "statut": result.get("statut_mahakim"),
                    "prochaine_audience": str(result.get("prochaine_audience") or ""),
                    "juge": result.get("juge"),
                    "card_info": result.get("card_info", {}),
                    "procedures": result.get("procedures", []),
                    "parties": result.get("parties", []),
                },
            })
        else:
            return JsonResponse({
                "ok": False,
                "message": result.get("error_message", "فشلت المزامنة"),
            })

    except ImportError:
        return JsonResponse({
            "ok": False,
            "message": "مكتبة Selenium غير مثبتة. قم بتثبيتها: pip install selenium",
        })
    except Exception as e:
        _mahakim_logger.exception("Erreur sync mahakim pour affaire %s", pk)
        return JsonResponse({
            "ok": False,
            "message": f"خطأ: {str(e)[:200]}",
        })


class MahakimSyncListView(SecureBase, NoPostOnReadOnlyMixin, ListView):
    """Page dédiée de synchronisation mahakim.ma."""
    model = Affaire
    template_name = "cabinet/mahakim_sync.html"
    permission_required = ()
    paginate_by = 50

    def get_queryset(self):
        return (
            Affaire.objects.filter(
                numero_dossier__isnull=False,
                code_categorie__isnull=False,
                annee_dossier__isnull=False,
            )
            .exclude(numero_dossier="")
            .exclude(annee_dossier="")
            .select_related("code_categorie", "juridiction", "juridiction__TribunalParent", "type_affaire", "statut_affaire")
            .order_by("-date_ouverture")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Ajouter le dernier sync pour chaque affaire
        affaire_ids = [a.pk for a in ctx["object_list"]]
        latest_syncs = {}
        for sync in MahakimSyncResult.objects.filter(affaire_id__in=affaire_ids).order_by("affaire_id", "-date_sync"):
            if sync.affaire_id not in latest_syncs:
                latest_syncs[sync.affaire_id] = sync
        # Build sync_data list for template
        sync_data = []
        for affaire in ctx["object_list"]:
            sync_data.append({
                "affaire": affaire,
                "sync": latest_syncs.get(affaire.pk),
            })
        ctx["sync_data"] = sync_data
        ctx["total_syncable"] = self.get_queryset().count()
        return ctx


def sync_all_mahakim(request):
    """Lance la synchronisation de toutes les affaires en arrière-plan."""
    if not request.user.is_authenticated:
        return JsonResponse({"ok": False, "message": "غير مسموح"}, status=403)

    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "POST فقط"}, status=405)

    from django.core.management import call_command

    def _run_sync():
        try:
            call_command("sync_mahakim")
        except Exception as e:
            _mahakim_logger.exception("Erreur sync_all_mahakim: %s", e)

    thread = threading.Thread(target=_run_sync, daemon=True)
    thread.start()

    return JsonResponse({
        "ok": True,
        "message": "بدأت المزامنة في الخلفية. أعد تحميل الصفحة بعد دقائق لرؤية النتائج.",
    })


@login_required
def fetch_mahakim_ids(request):
    """Lance la récupération des IDs des tribunaux en arrière-plan."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "POST فقط"}, status=405)

    task_id = str(_uuid.uuid4())

    with _mahakim_fetch_lock:
        _mahakim_fetch_tasks[task_id] = {
            "status": "running",
            "phase": "init",
            "current": 0,
            "total": 0,
            "name": "",
            "message": "جاري بدء المهمة...",
            "result": None,
            "errors": [],
        }

    def _run_fetch(tid):
        task = _mahakim_fetch_tasks[tid]
        try:
            from .services.mahakim_scraper import MahakimScraper

            def _progress(phase, current, total, name, message):
                with _mahakim_fetch_lock:
                    task["phase"] = phase
                    task["current"] = current
                    task["total"] = total
                    task["name"] = name
                    task["message"] = message

            with MahakimScraper(headless=False, timeout=120) as scraper:
                data = scraper.fetch_tribunal_ids(progress_callback=_progress)

            # Stocker les données brutes pour export Excel
            with _mahakim_fetch_lock:
                task["message"] = "جاري تحديث قاعدة البيانات..."
                task["phase"] = "saving"

            # Matching avec la BD
            matched = 0
            unmatched = []

            for item in data.get("appel", []):
                name = item.get("name", "").strip()
                mahakim_id = str(item.get("id", "")).strip()
                if not name or not mahakim_id:
                    continue
                qs = Juridiction.objects.filter(
                    nomtribunal_ar__icontains=name, TribunalParent__isnull=True
                )
                if not qs.exists():
                    qs = Juridiction.objects.filter(nomtribunal_ar__icontains=name)
                if qs.exists():
                    qs.update(id_mahakim=mahakim_id)
                    matched += qs.count()
                else:
                    unmatched.append(name)

            inserted = 0
            import re as _re

            def _normalize_ar(text):
                """Normalize Arabic: alef variants → ا, remove diacritics."""
                if not text:
                    return ''
                text = _re.sub('[إأآٱ]', 'ا', text)
                text = _re.sub('[\u064B-\u065F\u0670]', '', text)
                return _re.sub(r'\s+', ' ', text).strip()

            for item in data.get("premiere_instance", []):
                name = item.get("name", "").strip()
                mahakim_id = str(item.get("id", "")).strip()
                parent_name = item.get("parent_appel_name", "").strip()
                if not name or not mahakim_id:
                    continue

                # Resolve parent appeal court
                parent_jur = None
                if parent_name:
                    parent_jur = Juridiction.objects.filter(
                        nomtribunal_ar__icontains=parent_name,
                        TribunalParent__isnull=True,
                    ).first()

                # 1. Try exact icontains match (with parent)
                qs = Juridiction.objects.filter(
                    nomtribunal_ar__icontains=name, TribunalParent__isnull=False
                )
                if parent_name and qs.count() > 1:
                    qs_refined = qs.filter(TribunalParent__nomtribunal_ar__icontains=parent_name)
                    if qs_refined.exists():
                        qs = qs_refined

                # 2. Fallback: try without parent filter
                if not qs.exists():
                    qs = Juridiction.objects.filter(nomtribunal_ar__icontains=name)

                # 3. Fallback: normalized matching (handles alef/hamza differences)
                if not qs.exists():
                    norm_name = _normalize_ar(name)
                    for j in Juridiction.objects.filter(is_deleted=False).exclude(
                        nomtribunal_ar__icontains='استئناف'
                    ):
                        if _normalize_ar(j.nomtribunal_ar) == norm_name:
                            qs = Juridiction.objects.filter(pk=j.pk)
                            break

                if qs.exists():
                    # Update id_mahakim AND TribunalParent if missing
                    update_fields = {'id_mahakim': mahakim_id}
                    if parent_jur:
                        update_fields['TribunalParent'] = parent_jur
                    qs.filter(TribunalParent__isnull=True).update(**update_fields)
                    qs.filter(TribunalParent__isnull=False).update(id_mahakim=mahakim_id)
                    matched += qs.count()
                else:
                    # Auto-insert: create new Juridiction linked to parent cour d'appel
                    if parent_jur:
                        from .models import TypeJuridiction
                        type_pi = TypeJuridiction.objects.filter(code_type="TPI").first()
                        if type_pi:
                            Juridiction.objects.create(
                                code=mahakim_id,
                                nomtribunal_ar=name,
                                nomtribunal_fr=name,
                                type=type_pi,
                                TribunalParent=parent_jur,
                                id_mahakim=mahakim_id,
                            )
                            inserted += 1
                            matched += 1
                        else:
                            unmatched.append(name)
                    else:
                        unmatched.append(name)

            total_fetched = len(data.get("appel", [])) + len(data.get("premiere_instance", []))

            msg_parts = [f"تم جلب {total_fetched} محكمة", f"تم ربط {matched}"]
            if inserted:
                msg_parts.append(f"تم إنشاء {inserted} محكمة جديدة")
            if unmatched:
                msg_parts.append(f"لم يُطابَق {len(unmatched)}")

            with _mahakim_fetch_lock:
                task["status"] = "done"
                task["phase"] = "done"
                task["message"] = " — ".join(msg_parts)
                task["result"] = {
                    "appel": data.get("appel", []),
                    "premiere_instance": data.get("premiere_instance", []),
                    "matched": matched,
                    "inserted": inserted,
                    "unmatched": unmatched,
                    "total_fetched": total_fetched,
                }
                task["errors"] = data.get("errors", [])

        except ImportError:
            with _mahakim_fetch_lock:
                task["status"] = "error"
                task["message"] = "مكتبة Selenium غير مثبتة. قم بتثبيتها: pip install selenium"
        except Exception as e:
            _mahakim_logger.exception("Erreur fetch_mahakim_ids: %s", e)
            with _mahakim_fetch_lock:
                task["status"] = "error"
                task["message"] = f"خطأ: {str(e)[:200]}"

    thread = threading.Thread(target=_run_fetch, args=(task_id,), daemon=True)
    thread.start()

    return JsonResponse({"ok": True, "task_id": task_id})


@login_required
def fetch_mahakim_ids_status(request):
    """Retourne le statut de la tâche fetch_mahakim_ids (polling)."""
    task_id = request.GET.get("task_id", "")
    with _mahakim_fetch_lock:
        task = _mahakim_fetch_tasks.get(task_id)
    if not task:
        return JsonResponse({"ok": False, "message": "مهمة غير موجودة"}, status=404)

    with _mahakim_fetch_lock:
        return JsonResponse({
            "ok": True,
            "status": task["status"],
            "phase": task["phase"],
            "current": task["current"],
            "total": task["total"],
            "name": task["name"],
            "message": task["message"],
            "errors": task.get("errors", []),
            "result": task.get("result") if task["status"] == "done" else None,
        })


@login_required
def fetch_mahakim_ids_export(request):
    """Exporte les résultats du fetch en Excel."""
    task_id = request.GET.get("task_id", "")
    with _mahakim_fetch_lock:
        task = _mahakim_fetch_tasks.get(task_id)

    if not task or task["status"] != "done" or not task.get("result"):
        return JsonResponse({"ok": False, "message": "لا توجد نتائج للتصدير"}, status=404)

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return JsonResponse({"ok": False, "message": "مكتبة openpyxl غير مثبتة"}, status=500)

    result = task["result"]
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2C6B4F", end_color="2C6B4F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # --- Feuille 1: Cours d'appel ---
    ws_appel = wb.active
    ws_appel.title = "محاكم الاستئناف"
    ws_appel.sheet_view.rightToLeft = True
    headers_a = ["#", "معرف محاكم.ما", "اسم المحكمة"]
    for col, h in enumerate(headers_a, 1):
        cell = ws_appel.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for i, court in enumerate(result.get("appel", []), 1):
        ws_appel.cell(row=i + 1, column=1, value=i).border = thin_border
        ws_appel.cell(row=i + 1, column=2, value=court.get("id", "")).border = thin_border
        ws_appel.cell(row=i + 1, column=3, value=court.get("name", "")).border = thin_border
    ws_appel.column_dimensions["A"].width = 8
    ws_appel.column_dimensions["B"].width = 18
    ws_appel.column_dimensions["C"].width = 45

    # --- Feuille 2: Tribunaux de 1ère instance ---
    ws_pi = wb.create_sheet("المحاكم الابتدائية")
    ws_pi.sheet_view.rightToLeft = True
    headers_p = ["#", "معرف محاكم.ما", "اسم المحكمة", "محكمة الاستئناف الأم"]
    for col, h in enumerate(headers_p, 1):
        cell = ws_pi.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for i, court in enumerate(result.get("premiere_instance", []), 1):
        ws_pi.cell(row=i + 1, column=1, value=i).border = thin_border
        ws_pi.cell(row=i + 1, column=2, value=court.get("id", "")).border = thin_border
        ws_pi.cell(row=i + 1, column=3, value=court.get("name", "")).border = thin_border
        ws_pi.cell(row=i + 1, column=4, value=court.get("parent_appel_name", "")).border = thin_border
    ws_pi.column_dimensions["A"].width = 8
    ws_pi.column_dimensions["B"].width = 18
    ws_pi.column_dimensions["C"].width = 45
    ws_pi.column_dimensions["D"].width = 40

    # --- Feuille 3: Non-appariés ---
    unmatched = result.get("unmatched", [])
    if unmatched:
        ws_un = wb.create_sheet("غير مطابقة")
        ws_un.sheet_view.rightToLeft = True
        headers_u = ["#", "اسم المحكمة"]
        for col, h in enumerate(headers_u, 1):
            cell = ws_un.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
            cell.alignment = header_align
            cell.border = thin_border
        for i, name in enumerate(unmatched, 1):
            ws_un.cell(row=i + 1, column=1, value=i).border = thin_border
            ws_un.cell(row=i + 1, column=2, value=name).border = thin_border
        ws_un.column_dimensions["A"].width = 8
        ws_un.column_dimensions["B"].width = 50

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="mahakim_tribunaux.xlsx"'
    return response


# =============================================================
# MAHAKIM — SESSION SYNC (مزامنة جدول الجلسات)
# =============================================================

@login_required
def mahakim_preview_sessions(request):
    """Returns modal HTML for session sync with court selector + date picker."""
    # Appeal courts for the main dropdown
    appel = Juridiction.objects.filter(
        nomtribunal_ar__icontains='استئناف',
        is_deleted=False,
    ).select_related("type").order_by("nomtribunal_ar")

    html = render_to_string("modals/_mahakim_preview_sessions.html", {
        "juridictions_appel": appel,
    }, request=request)
    return HttpResponse(html)


@login_required
def get_tribunaux_pi(request):
    """API: returns PI tribunals for a given appeal court (parent)."""
    parent_id = request.GET.get("parent_id")
    if not parent_id:
        return JsonResponse({"tribunaux": []})

    tribunaux = Juridiction.objects.filter(
        TribunalParent_id=parent_id,
        is_deleted=False,
    ).order_by("nomtribunal_ar").values("pk", "nomtribunal_ar", "id_mahakim")

    return JsonResponse({
        "tribunaux": [
            {"pk": str(t["pk"]), "nom": t["nomtribunal_ar"], "id_mahakim": t["id_mahakim"] or ""}
            for t in tribunaux
        ]
    })


@login_required
def sync_sessions_mahakim(request):
    """Sync session schedule from mahakim.ma for a given court + date."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "POST فقط"}, status=405)

    import json as _j
    try:
        body = _j.loads(request.body)
    except (ValueError, TypeError):
        body = request.POST

    juridiction_id = body.get("juridiction_id")
    date_seance = body.get("date")
    type_seance = body.get("type_seance") or None

    if not juridiction_id or not date_seance:
        return JsonResponse({
            "ok": False,
            "message": "يجب اختيار المحكمة والتاريخ"
        })

    try:
        juridiction = Juridiction.objects.get(pk=juridiction_id)
    except Juridiction.DoesNotExist:
        return JsonResponse({"ok": False, "message": "المحكمة غير موجودة"})

    try:
        from .services.mahakim_scraper import MahakimScraper

        # Auto-resolve parent appeal court for first-instance tribunals
        is_premiere = bool(juridiction.TribunalParent)
        nom_tribunal = juridiction.nomtribunal_ar

        if is_premiere:
            # First-instance court → use parent appeal court for scraping
            parent = juridiction.TribunalParent
            nom_tribunal_appel = parent.nomtribunal_ar
            id_mahakim = parent.id_mahakim or juridiction.id_mahakim or None
        else:
            # Already an appeal court (or CC)
            nom_tribunal_appel = None
            id_mahakim = juridiction.id_mahakim or None

        with MahakimScraper(headless=False, timeout=30) as scraper:
            result = scraper.scrape_sessions(
                id_mahakim_tribunal=id_mahakim,
                date_seance=date_seance,
                type_seance=type_seance,
                is_premiere_instance=is_premiere,
                nom_tribunal=nom_tribunal,
                nom_tribunal_appel=nom_tribunal_appel,
            )

        # Store sync result
        MahakimSyncResult.objects.create(
            sync_type='sessions',
            statut_mahakim=f"جلسات {juridiction.nomtribunal_ar} — {date_seance}",
            observations=f"{len(result.get('sessions', []))} جلسة",
            raw_html=(result.get("raw_html") or "")[:50000],
            success=result.get("success", False),
            error_message=result.get("error_message"),
            procedures_json=result.get("sessions"),
        )

        if result["success"]:
            return JsonResponse({
                "ok": True,
                "message": f"تم جلب {len(result.get('sessions', []))} جلسة",
                "sessions": result.get("sessions", []),
            })
        else:
            return JsonResponse({
                "ok": False,
                "message": result.get("error_message", "فشلت المزامنة"),
            })

    except ImportError:
        return JsonResponse({
            "ok": False,
            "message": "مكتبة Selenium غير مثبتة. قم بتثبيتها: pip install selenium",
        })
    except Exception as e:
        _mahakim_logger.exception("Erreur sync sessions mahakim")
        return JsonResponse({
            "ok": False,
            "message": f"خطأ: {str(e)[:200]}",
        })


# =============================================================
# المسطرة الغيابية — CONTUMACE SYNC
# =============================================================

_contumace_sync_lock = threading.Lock()
_contumace_sync_tasks = {}  # task_id → {status, phase, current_page, total_pages, records_count, message}


class ContumaceListView(SecureBase, SearchListMixin, ListView):
    model = ContumaceRecord
    template_name = "cabinet/contumace_list.html"
    context_object_name = "records"
    paginate_by = 50
    permission_required = ''
    search_fields = ["nom_accuse", "numero_dossier", "cour_appel", "numero_carte", "nom_pere", "nom_mere"]

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["total_records"] = ContumaceRecord.objects.count()
        return ctx


@login_required
def mahakim_preview_contumace(request):
    """Returns modal HTML for contumace sync explanation + optional search."""
    html = render_to_string("modals/_mahakim_preview_contumace.html", {}, request=request)
    return HttpResponse(html)


@login_required
def sync_contumace_mahakim(request):
    """Lance le scraping de la page contumace en arrière-plan."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "POST فقط"}, status=405)

    import json as _j
    try:
        body = _j.loads(request.body)
    except (ValueError, TypeError):
        body = request.POST

    search_query = (body.get("search_query") or "").strip() or None

    task_id = str(_uuid.uuid4())

    with _contumace_sync_lock:
        _contumace_sync_tasks[task_id] = {
            "status": "running",
            "phase": "init",
            "current_page": 0,
            "total_pages": 0,
            "records_count": 0,
            "message": "جاري بدء المهمة...",
        }

    def _run_contumace(tid):
        task = _contumace_sync_tasks[tid]
        scraper = None
        try:
            from .services.mahakim_scraper import MahakimScraper

            def _progress(phase, current_page, total_pages, records_count, message):
                with _contumace_sync_lock:
                    task["phase"] = phase
                    task["current_page"] = current_page
                    task["total_pages"] = total_pages
                    task["records_count"] = records_count
                    task["message"] = message

            # Load existing (numero_dossier, cour_appel) keys from DB
            # so the scraper can skip records that are already stored
            existing_keys = set(
                ContumaceRecord.objects.values_list('numero_dossier', 'cour_appel')
            )

            # Don't use 'with' — on error we keep the browser open for debugging
            scraper = MahakimScraper(headless=False, timeout=60)
            scraper._create_driver()
            data = scraper.scrape_contumace(
                search_query=search_query,
                progress_callback=_progress,
                existing_keys=existing_keys,
            )

            if data["success"]:
                # Success: close browser, save records
                scraper.close()

                with _contumace_sync_lock:
                    task["phase"] = "saving"
                    task["message"] = "جاري حفظ السجلات في قاعدة البيانات..."

                saved = 0
                for rec in data.get("records", []):
                    if not rec.get("numero_dossier"):
                        continue
                    ContumaceRecord.objects.update_or_create(
                        numero_dossier=rec["numero_dossier"],
                        cour_appel=rec["cour_appel"],
                        defaults={
                            "nom_accuse": rec.get("nom_accuse", ""),
                            "nom_pere": rec.get("nom_pere", ""),
                            "nom_mere": rec.get("nom_mere", ""),
                            "numero_carte": rec.get("numero_carte", ""),
                            "details_text": rec.get("details_text", ""),
                        }
                    )
                    saved += 1

                with _contumace_sync_lock:
                    task["status"] = "done"
                    task["phase"] = "done"
                    task["records_count"] = saved
                    task["message"] = f"تم جلب وحفظ {saved} سجل بنجاح"
            else:
                # Error from scraper: keep browser open for debugging
                with _contumace_sync_lock:
                    task["status"] = "error"
                    task["message"] = (
                        (data.get("error_message") or "فشلت المزامنة")
                        + " — المتصفح مفتوح للتشخيص"
                    )

        except ImportError:
            with _contumace_sync_lock:
                task["status"] = "error"
                task["message"] = "مكتبة Selenium غير مثبتة. قم بتثبيتها: pip install selenium"
        except Exception as e:
            _mahakim_logger.exception("Erreur sync contumace: %s", e)
            # Keep browser open on unexpected error
            with _contumace_sync_lock:
                task["status"] = "error"
                task["message"] = f"خطأ: {str(e)[:200]} — المتصفح مفتوح للتشخيص"

    thread = threading.Thread(target=_run_contumace, args=(task_id,), daemon=True)
    thread.start()

    return JsonResponse({"ok": True, "task_id": task_id})


@login_required
def contumace_sync_status(request):
    """Retourne le statut de la tâche sync contumace (polling)."""
    task_id = request.GET.get("task_id", "")
    with _contumace_sync_lock:
        task = _contumace_sync_tasks.get(task_id)

    if not task:
        return JsonResponse({"ok": False, "message": "مهمة غير موجودة"})

    with _contumace_sync_lock:
        return JsonResponse({
            "ok": True,
            "status": task["status"],
            "phase": task["phase"],
            "current_page": task["current_page"],
            "total_pages": task["total_pages"],
            "records_count": task["records_count"],
            "message": task["message"],
        })


# =============================================================
# حاسبة الرسوم القضائية — TAX CALCULATOR (محلي)
# ملحق المرسوم رقم 2.58.1151 — المصاريف القضائية
# =============================================================


class TaxCalculatorView(SecureBase, TemplateView):
    template_name = "cabinet/tax_calculator.html"
    permission_required = ''


# =============================================================
# DOCUMENT VIEWER (معاينة المرفقات)
# =============================================================

# Registry: model_name → (Model, file_field, permission)
_FILE_REGISTRY = {
    "piecejointe": (PieceJointe, "fichier", "cabinet.view_piecejointe"),
    "depense": (Depense, "piece", "cabinet.view_depense"),
    "recette": (Recette, "piece", "cabinet.view_recette"),
    "expertise": (Expertise, "rapport", "cabinet.view_expertise"),
    "notification": (Notification, "preuve", "cabinet.view_notification"),
    "avertissement": (Avertissement, "preuve_envoi", "cabinet.view_avertissement"),
}

_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".svg"}
_PDF_EXTS = {".pdf"}


def _detect_doc_type(filename):
    import os
    ext = os.path.splitext(filename)[1].lower()
    if ext in _PDF_EXTS:
        return "pdf"
    if ext in _IMAGE_EXTS:
        return "image"
    return "other"


@ensure_csrf_cookie
def document_viewer(request):
    """Vue HTMX qui renvoie le contenu modal pour visionner un document."""
    if not request.user.is_authenticated:
        return HttpResponse("غير مسموح", status=403)

    model_name = request.GET.get("model", "").lower()
    pk = request.GET.get("pk")
    field_name = request.GET.get("field")  # optionnel

    if model_name not in _FILE_REGISTRY:
        return HttpResponse("نموذج غير معروف", status=400)

    Model, default_field, perm = _FILE_REGISTRY[model_name]

    if not request.user.has_perm(perm):
        return HttpResponse("غير مسموح", status=403)

    obj = get_object_or_404(Model, pk=pk)
    field = field_name or default_field
    file_obj = getattr(obj, field, None)

    if not file_obj:
        return HttpResponse("لا يوجد ملف مرفق", status=404)

    file_url = file_obj.url
    doc_type = _detect_doc_type(file_obj.name)
    title = str(obj)

    html = render_to_string("modals/_document_viewer.html", {
        "title": title,
        "file_url": file_url,
        "doc_type": doc_type,
    }, request=request)
    return HttpResponse(html)


# =============================================================
# PRINT DOCUMENTS (طباعة المستندات)
# =============================================================

def print_documents(request):
    """Page autonome d'impression de documents filtrés."""
    if not request.user.is_authenticated:
        return HttpResponse("غير مسموح", status=403)

    source = request.GET.get("source", "").lower()

    SOURCE_MAP = {
        "depense": (Depense, "piece", "cabinet.view_depense", DepenseFilter),
        "recette": (Recette, "piece", "cabinet.view_recette", RecetteFilter),
        "piecejointe": (PieceJointe, "fichier", "cabinet.view_piecejointe", PieceJointeFilter),
    }

    if source not in SOURCE_MAP:
        return HttpResponse("مصدر غير صالح", status=400)

    Model, file_field, perm, FilterClass = SOURCE_MAP[source]

    if not request.user.has_perm(perm):
        return HttpResponse("غير مسموح", status=403)

    qs = Model.objects.all()
    fs = FilterClass(request.GET, queryset=qs)
    qs = fs.qs[:100]  # max 100 documents

    documents = []
    for obj in qs:
        f = getattr(obj, file_field, None)
        if f:
            doc_type = _detect_doc_type(f.name)
            documents.append({
                "title": str(obj),
                "url": f.url,
                "type": doc_type,
            })

    return render(request, "print/documents.html", {
        "documents": documents,
        "source_label": {"depense": "المصاريف", "recette": "الإيرادات", "piecejointe": "المرفقات"}.get(source, ""),
        "count": len(documents),
    })


# ====== دليل الاستعمال ======
@login_required
def user_guide(request):
    return render(request, "guide/user_guide.html", {"today": timezone.localdate()})


# =============================================================
# WHATSAPP (Twilio) — Templates + Journal + Envoi manuel
# =============================================================

class WhatsAppTemplateList(SecureBase, ListView):
    model = WhatsAppTemplate
    template_name = "whatsapp/template_list.html"
    context_object_name = "templates"
    permission_required = "cabinet.view_affaire"
    paginate_by = 20


class WhatsAppTemplateCreate(UIPermRequiredMixin, SecureBase, CreateView):
    ui_perm = "ui_btn_add"
    model = WhatsAppTemplate
    template_name = "whatsapp/template_form.html"
    fields = ["nom", "kind", "body", "is_active", "twilio_content_sid"]
    permission_required = "cabinet.change_affaire"
    success_url = reverse_lazy("cabinet:whatsapp_template_list")


class WhatsAppTemplateUpdate(UIPermRequiredMixin, SecureBase, UpdateView):
    ui_perm = "ui_btn_edit"
    model = WhatsAppTemplate
    template_name = "whatsapp/template_form.html"
    fields = ["nom", "kind", "body", "is_active", "twilio_content_sid"]
    permission_required = "cabinet.change_affaire"
    success_url = reverse_lazy("cabinet:whatsapp_template_list")


class WhatsAppTemplateDelete(UIPermRequiredMixin, SecureBase, DeleteView):
    ui_perm = "ui_btn_delete"
    model = WhatsAppTemplate
    template_name = "whatsapp/template_confirm_delete.html"
    permission_required = "cabinet.change_affaire"
    success_url = reverse_lazy("cabinet:whatsapp_template_list")


class WhatsAppMessageList(SecureBase, ListView):
    model = WhatsAppMessage
    template_name = "whatsapp/message_list.html"
    context_object_name = "messages_qs"
    permission_required = "cabinet.view_affaire"
    paginate_by = 30

    def get_queryset(self):
        return (
            WhatsAppMessage.objects.select_related("affaire", "audience", "template")
            .order_by("-created_at")
        )


from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST


@login_required(login_url=reverse_lazy('authui:login'))
def portail_issue_link(request: HttpRequest, partie_id) -> HttpResponse:
    """Action back-office: générer et envoyer un magic link à une partie."""
    if request.method != "POST":
        return HttpResponse("Method not allowed", status=405)
    partie = get_object_or_404(Partie, pk=partie_id)
    from .services.portail_auth import issue_token, send_magic_link
    tok = issue_token(partie, request=request)
    base_url = f"{request.scheme}://{request.get_host()}"
    sent = send_magic_link(partie, tok.token, base_url=base_url)
    channels = []
    if sent.get("email"):
        channels.append("بريد إلكتروني")
    if sent.get("whatsapp"):
        channels.append("واتساب")
    if channels:
        messages.success(request, f"تم إرسال رابط البوابة عبر: {', '.join(channels)}.")
    else:
        messages.warning(request, f"تم إنشاء الرابط لكن لم يُرسل (لا بريد/هاتف). الرابط: {sent.get('url')}")
    return redirect(request.POST.get("next") or reverse("cabinet:partie_detail", kwargs={"pk": partie.pk}))


@csrf_exempt
@require_POST
def whatsapp_webhook(request: HttpRequest) -> HttpResponse:
    """Webhook Twilio pour les messages WhatsApp entrants.

    Twilio envoie un POST x-www-form-urlencoded contenant entre autres:
    From, Body, ProfileName, MessageSid, NumMedia.

    Réponse TwiML XML pour que Twilio renvoie le message à l'utilisateur.
    """
    from_number = request.POST.get("From", "")
    body = request.POST.get("Body", "")
    profile_name = request.POST.get("ProfileName", "")
    message_sid = request.POST.get("MessageSid", "")

    # Validation optionnelle de signature Twilio (production)
    try:
        from .services.twilio_security import validate_twilio_signature
        if not validate_twilio_signature(request):
            return HttpResponse("Invalid signature", status=403)
    except Exception:
        # En dev / sans signature → on accepte
        pass

    from .services.whatsapp_bot import handle_inbound_message
    try:
        reply = handle_inbound_message(from_number, body,
                                       profile_name=profile_name,
                                       message_sid=message_sid)
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception("WhatsApp bot failed")
        reply = "⚠️ حدث خطأ مؤقت. حاول مجددا لاحقًا."

    # Réponse TwiML
    from xml.sax.saxutils import escape
    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Response>'
        f'<Message>{escape(reply)}</Message>'
        '</Response>'
    )
    return HttpResponse(xml, content_type="application/xml")


@login_required(login_url=reverse_lazy('authui:login'))
def whatsapp_send_audience_reminder(request: HttpRequest, audience_id) -> HttpResponse:
    """Envoi manuel d'un rappel WhatsApp pour une audience donnée."""
    if request.method != "POST":
        return HttpResponse("Method not allowed", status=405)

    audience = get_object_or_404(Audience.objects.select_related("affaire", "affaire__juridiction"), pk=audience_id)
    from .services.twilio_client import send_audience_reminder
    msg = send_audience_reminder(audience, manual=True)

    if msg and msg.status in ("sent", "dry_run"):
        messages.success(request, f"تم إرسال التذكير عبر واتساب ({msg.get_status_display()}).")
    else:
        err = (msg.error_message if msg else "خطأ غير معروف") or "خطأ"
        messages.warning(request, f"فشل الإرسال: {err}")

    next_url = request.POST.get("next") or reverse("cabinet:audience_detail", kwargs={"pk": audience.pk})
    return redirect(next_url)


# =============================================================
# Recherche jurisprudentielle sémantique
# =============================================================

@login_required(login_url=reverse_lazy('authui:login'))
def jurisprudence_search(request: HttpRequest) -> HttpResponse:
    """Recherche sémantique sur les analyses de décisions (embeddings)."""
    from .services.embeddings import search_decisions
    from .models import DecisionAnalysis

    q = (request.GET.get("q") or "").strip()
    results = []
    if q and len(q) >= 3:
        scored = search_decisions(q, top_k=10)
        for analysis, score in scored:
            results.append({
                "analysis": analysis,
                "decision": analysis.decision,
                "affaire": analysis.decision.affaire,
                "score": round(score, 4),
                "score_pct": int(round(score * 100)),
            })

    total_indexed = DecisionAnalysis.objects.exclude(embedding__isnull=True).count()
    total_decisions = DecisionAnalysis.objects.count()

    return render(request, "cabinet/jurisprudence_search.html", {
        "q": q,
        "results": results,
        "total_indexed": total_indexed,
        "total_decisions": total_decisions,
    })


# =============================================================
# Carte des juridictions (Leaflet + OpenStreetMap)
# =============================================================

@login_required(login_url=reverse_lazy('authui:login'))
def juridictions_map(request: HttpRequest) -> HttpResponse:
    """Page de carte affichant les juridictions géolocalisées."""
    from django.db.models import Count

    qs = (
        Juridiction.objects
        .annotate(affaires_count=Count("affaire", distinct=True))
        .order_by("villetribunal_ar")
    )

    points = []
    missing = []
    for j in qs:
        if j.latitude is not None and j.longitude is not None:
            points.append({
                "id": j.pk,
                "nom_ar": j.nomtribunal_ar or "",
                "nom_fr": j.nomtribunal_fr or "",
                "ville": j.villetribunal_ar or j.villetribunal_fr or "",
                "lat": float(j.latitude),
                "lng": float(j.longitude),
                "telephone": j.telephonetribunal or "",
                "affaires_count": j.affaires_count,
                "url_detail": j.get_absolute_url(),
                "url_affaires": reverse("cabinet:affaire_list") + f"?juridiction={j.pk}",
                "url_directions": j.google_maps_directions_url,
            })
        else:
            missing.append(j)

    return render(request, "cabinet/juridictions_map.html", {
        "points": points,
        "missing": missing,
        "missing_count": len(missing),
        "total_count": len(points) + len(missing),
    })


# =============================================================
# AI — Analyse de décision via Claude
# =============================================================

@login_required(login_url=reverse_lazy('authui:login'))
def decision_ai_analyze(request: HttpRequest, pk) -> HttpResponse:
    """Lance une analyse IA sur une Decision (texte saisi ou PDF joint)."""
    if request.method != "POST":
        return HttpResponse("Method not allowed", status=405)

    decision = get_object_or_404(Decision.objects.select_related("affaire"), pk=pk)
    source_text = (request.POST.get("source_text") or "").strip()
    piece_id = request.POST.get("piece_jointe_id") or None

    piece_jointe = None
    if piece_id:
        try:
            piece_jointe = PieceJointe.objects.filter(pk=piece_id, affaire=decision.affaire).first()
        except Exception:
            piece_jointe = None

    from .services.ai_client import analyze_decision
    try:
        analysis = analyze_decision(decision, source_text=source_text or None, piece_jointe=piece_jointe)
    except Exception as e:
        messages.error(request, f"فشل التحليل: {e}")
        return redirect(reverse("cabinet:decision_detail", kwargs={"pk": pk}))

    if analysis.error_message:
        messages.warning(request, f"تم الحفظ مع خطأ: {analysis.error_message[:200]}")
    elif analysis.is_dry_run:
        messages.info(request, "تمت المحاكاة بنجاح (لم يتم تكوين مفتاح Anthropic).")
    else:
        messages.success(request, f"تم التحليل بنجاح بواسطة {analysis.model_used}.")

    return redirect(reverse("cabinet:decision_detail", kwargs={"pk": pk}))
